import os
import re
import json
import time
import mimetypes
import threading
import schedule
import httpx
import psycopg2
from flask import Flask, request, jsonify, Response
from datetime import datetime, timedelta

import nk_integration as nk  # интеграция с Национальным каталогом (Честный ЗНАК)

app = Flask(__name__)

# ===================== КОНФИГ (переменные окружения) =====================

WB_API_TOKEN = os.environ.get("WB_API_TOKEN", "").strip()
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()

# Bitrix24 — локальное приложение (OAuth)
BITRIX_CLIENT_ID = os.environ.get("BITRIX_CLIENT_ID", "").strip()
BITRIX_CLIENT_SECRET = os.environ.get("BITRIX_CLIENT_SECRET", "").strip()
BITRIX_APP_TOKEN = os.environ.get("BITRIX_APP_TOKEN", "").strip()
PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").strip().rstrip("/")
# Домен портала Битрикс24 для REST-вызовов. Нужен как запасной, если в OAuth
# случайно сохранился oauth.bitrix.info (его Битрикс возвращает в ответе refresh).
BITRIX_PORTAL_DOMAIN = os.environ.get("BITRIX_PORTAL_DOMAIN", "joto.bitrix24.ru").strip()

# Кому слать отчёт по задачам и алерты CTR
REPORT_USER_ID = os.environ.get("REPORT_USER_ID", "226").strip()
CTR_ALERT_DIALOG = os.environ.get("CTR_ALERT_DIALOG", "chat2024").strip()
# Чат «Отдел продаж» в Битрикс24 для отчёта по сезонной распродаже (DIALOG_ID: chatXXXX или ID пользователя)
# По умолчанию — chat2024 (https://joto.bitrix24.ru/online/?IM_DIALOG=chat2024)
SALES_DEPT_DIALOG = os.environ.get("SALES_DEPT_DIALOG", "").strip() or "chat2024"
# Категория для еженедельного авто-отчёта по сезонной распродаже и конец сезона
SEASON_REPORT_CATEGORY = os.environ.get("SEASON_REPORT_CATEGORY", "10").strip()
SEASON_END_DATE = os.environ.get("SEASON_END_DATE", "2026-08-31").strip()
# Получатель отчёта по распродаже. Можно указать имя сотрудника (ищется в Битриксе),
# ID пользователя (число) или чат (chatXXXX). По умолчанию — лично Татьяне.
SEASON_REPORT_TO = os.environ.get("SEASON_REPORT_TO", "Татьяна").strip()
# Маржа: себестоимость как доля от обычной цены (0..1) и минимальная наценка к ней.
# Ограничивают максимально возможную скидку, чтобы распродажа не уходила ниже маржи.
try:
    SEASON_COST_SHARE = float(os.environ.get("SEASON_COST_SHARE", "0.35") or 0.35)
except Exception:
    SEASON_COST_SHARE = 0.35
try:
    SEASON_MIN_MARGIN = float(os.environ.get("SEASON_MIN_MARGIN", "0.10") or 0.10)
except Exception:
    SEASON_MIN_MARGIN = 0.10
# Период для расчёта % выкупа (выкупы отстают от заказов — берём «отстоявшееся» окно).
try:
    SEASON_BUYOUT_DAYS = int(os.environ.get("SEASON_BUYOUT_DAYS", "60") or 60)
except Exception:
    SEASON_BUYOUT_DAYS = 60

def is_silent_dialog(dialog_id):
    """Чаты только для отчётов/алертов: бот туда пишет сам, но не приветствует
    и не ведёт диалог по созданию артикулов (например, чат отдела продаж)."""
    d = str(dialog_id or "").strip()
    silent = {str(SALES_DEPT_DIALOG).strip(), str(CTR_ALERT_DIALOG).strip()}
    silent.discard("")
    return d in silent

# Ссылка на логотип (необязательно). Приоритет выше файла logo.* в репозитории.
LOGO_URL = os.environ.get("LOGO_URL", "").strip()

# Имя бота, которое увидят пользователи в Битриксе
BOT_NAME = "Article Generator"
BOT_CODE = "joto_article_bot"

EVENT_HANDLER_URL = f"{PUBLIC_BASE_URL}/bitrix/events" if PUBLIC_BASE_URL else ""
# Пункты левого меню Bitrix24 — каждый раздел отдельным пунктом
LEFT_MENU_HANDLER_URL = f"{PUBLIC_BASE_URL}/cards" if PUBLIC_BASE_URL else ""
LEFT_MENU_TITLE = "Карточки WB"
# Отдельный пункт левого меню → раздел распродажи сезона (отчёт по шортам)
SEASON_MENU_HANDLER_URL = f"{PUBLIC_BASE_URL}/season" if PUBLIC_BASE_URL else ""
SEASON_MENU_TITLE = "Распродажа сезона"

# Встроенные категории по инструкции JOTO (включая словоформы для чата)
CATEGORIES = {
    "жилет": "01", "жилеты": "01",
    "куртка": "02", "куртки": "02",
    "водолазка": "03", "водолазки": "03",
    "джинсы": "04",
    "худи": "05",
    "свитер": "06", "свитера": "06",
    "лонгслив": "07", "лонгсливы": "07",
    "брюки": "09",
    "шорты": "10",
    "футболка": "11", "футболки": "11",
}

CATS_LIST = "жилет, куртка, водолазка, джинсы, худи, свитер, лонгслив, брюки, шорты, футболка"

user_states = {}

# ===================== БД =====================

def get_db():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS model_counters (
                category_code VARCHAR(10) PRIMARY KEY,
                counter INTEGER DEFAULT 0
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS bitrix_oauth (
                id           INTEGER PRIMARY KEY,
                access_token TEXT,
                refresh_token TEXT,
                expires_at   BIGINT,
                domain       TEXT,
                member_id    TEXT,
                bot_id       TEXT,
                app_token    TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                name TEXT PRIMARY KEY,
                code VARCHAR(2) NOT NULL UNIQUE
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        print("БД инициализирована")
    except Exception as e:
        print(f"Ошибка БД: {e}")

# ===================== КАТЕГОРИИ (встроенные + добавленные) =====================

CATEGORY_TITLES = {
    "жилет": "Жилет", "куртка": "Куртка", "водолазка": "Водолазка",
    "джинсы": "Джинсы", "худи": "Худи", "свитер": "Свитер",
    "лонгслив": "Лонгслив", "брюки": "Брюки", "шорты": "Шорты", "футболка": "Футболка",
}

def db_list_categories():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT name, code FROM categories ORDER BY code")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [(r[0], r[1]) for r in rows]
    except Exception as e:
        print(f"Ошибка db_list_categories: {e}")
        return []

def resolve_category_code(name):
    name = (name or "").strip().lower()
    if name in CATEGORIES:
        return CATEGORIES[name]
    for n, code in db_list_categories():
        if n.lower() == name:
            return code
    return None

def all_categories():
    seen_codes = set()
    items = []
    for value, title in CATEGORY_TITLES.items():
        code = CATEGORIES.get(value)
        if code and code not in seen_codes:
            items.append({"value": value, "title": title, "code": code})
            seen_codes.add(code)
    for name, code in db_list_categories():
        if code not in seen_codes:
            items.append({"value": name, "title": name.capitalize(), "code": code})
            seen_codes.add(code)
    items.sort(key=lambda x: x["code"])
    return items

def used_category_codes():
    codes = set(CATEGORIES.values())
    for _, code in db_list_categories():
        codes.add(code)
    return codes

def next_free_category_code():
    used = used_category_codes()
    for i in range(1, 100):
        c = str(i).zfill(2)
        if c not in used:
            return c
    return None

def add_category(name, code=None):
    name = (name or "").strip().lower()
    if not name:
        return False, "Укажите название категории"
    if resolve_category_code(name):
        return False, "Такая категория уже есть"
    if code:
        code = str(code).strip()
        if not re.fullmatch(r"\d{2}", code):
            return False, "Код должен быть из 2 цифр"
        if code in used_category_codes():
            return False, "Этот код уже занят"
    else:
        code = next_free_category_code()
        if not code:
            return False, "Свободных кодов не осталось"
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("INSERT INTO categories (name, code) VALUES (%s, %s)", (name, code))
        conn.commit()
        cur.close()
        conn.close()
        return True, code
    except Exception as e:
        print(f"Ошибка add_category: {e}")
        return False, "Ошибка сохранения"

# ===================== СЧЁТЧИКИ НОМЕРОВ =====================

def get_next_model_number(category_code):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO model_counters (category_code, counter)
            VALUES (%s, 1)
            ON CONFLICT (category_code) DO UPDATE
            SET counter = model_counters.counter + 1
            RETURNING counter
        """, (category_code,))
        number = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        return str(number).zfill(3)
    except Exception as e:
        print(f"Ошибка счётчика: {e}")
        return "001"

def get_current_counter(category_code):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT counter FROM model_counters WHERE category_code=%s", (category_code,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return row[0] if row else 0
    except Exception as e:
        return 0

def set_counter(category_code, value):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO model_counters (category_code, counter)
            VALUES (%s, %s)
            ON CONFLICT (category_code) DO UPDATE SET counter = EXCLUDED.counter
        """, (category_code, value))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Ошибка set_counter: {e}")

# ===================== СИНХРОНИЗАЦИЯ С WB (уникальность артикулов) =====================

ARTICLE_RE = re.compile(r"^J(\d{2})(\d{3})", re.IGNORECASE)

_wb_used_cache = {"ts": 0.0, "used": {}}
_WB_CACHE_TTL = 300

def fetch_wb_vendor_codes():
    codes = []
    if not WB_API_TOKEN:
        return codes
    url = "https://content-api.wildberries.ru/content/v2/get/cards/list"
    headers = {"Authorization": WB_API_TOKEN}
    cursor = {"limit": 100}
    for _ in range(200):
        payload = {"settings": {"cursor": cursor, "filter": {"withPhoto": -1}}}
        resp = httpx.post(url, headers=headers, json=payload, timeout=30)
        if resp.status_code != 200:
            print(f"WB Content API {resp.status_code}: {resp.text[:200]}")
            break
        data = resp.json()
        cards = data.get("cards", []) or []
        for c in cards:
            vc = c.get("vendorCode")
            if vc:
                codes.append(vc)
        cur = data.get("cursor", {}) or {}
        total = cur.get("total", 0)
        if total < cursor["limit"]:
            break
        cursor = {"limit": 100, "updatedAt": cur.get("updatedAt"), "nmID": cur.get("nmID")}
    return codes

def _used_numbers_by_code(codes):
    used = {}
    for vc in codes:
        m = ARTICLE_RE.match((vc or "").strip())
        if m:
            cc = m.group(1)
            nn = int(m.group(2))
            used.setdefault(cc, set()).add(nn)
    return used

def get_used_numbers(category_code):
    now = time.time()
    if now - _wb_used_cache["ts"] > _WB_CACHE_TTL or not _wb_used_cache["used"]:
        try:
            codes = fetch_wb_vendor_codes()
            if codes:
                _wb_used_cache["used"] = _used_numbers_by_code(codes)
                _wb_used_cache["ts"] = now
                print(f"WB: загружено артикулов {len(codes)}")
        except Exception as e:
            print(f"Ошибка синхронизации с WB: {e}")
    return _wb_used_cache["used"].get(category_code, set())

def peek_next_number(category_code):
    used = get_used_numbers(category_code)
    base = max([get_current_counter(category_code)] + (list(used) or [0]))
    n = base + 1
    while n in used:
        n += 1
    return str(n).zfill(3)

def reserve_next_number(category_code):
    used = get_used_numbers(category_code)
    base = max([get_current_counter(category_code)] + (list(used) or [0]))
    n = base + 1
    while n in used:
        n += 1
    set_counter(category_code, n)
    return str(n).zfill(3)

# ===================== OAuth-хранилище (Postgres) =====================

def save_oauth(access_token, refresh_token, expires_in, domain,
               member_id=None, bot_id=None, app_token=None):
    try:
        expires_at = int(time.time()) + int(expires_in or 3600) - 60
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bitrix_oauth
                (id, access_token, refresh_token, expires_at, domain, member_id, bot_id, app_token)
            VALUES (1, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO UPDATE SET
                access_token  = EXCLUDED.access_token,
                refresh_token = EXCLUDED.refresh_token,
                expires_at    = EXCLUDED.expires_at,
                domain        = EXCLUDED.domain,
                member_id     = COALESCE(EXCLUDED.member_id, bitrix_oauth.member_id),
                bot_id        = COALESCE(EXCLUDED.bot_id, bitrix_oauth.bot_id),
                app_token     = COALESCE(EXCLUDED.app_token, bitrix_oauth.app_token)
        """, (access_token, refresh_token, expires_at, domain, member_id, bot_id, app_token))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Ошибка save_oauth: {e}")

def load_oauth():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT access_token, refresh_token, expires_at, domain, member_id, bot_id, app_token
            FROM bitrix_oauth WHERE id=1
        """)
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return None
        return {
            "access_token": row[0],
            "refresh_token": row[1],
            "expires_at": row[2] or 0,
            "domain": row[3],
            "member_id": row[4],
            "bot_id": row[5],
            "app_token": row[6],
        }
    except Exception as e:
        print(f"Ошибка load_oauth: {e}")
        return None

def set_bot_id(bot_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE bitrix_oauth SET bot_id=%s WHERE id=1", (str(bot_id),))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Ошибка set_bot_id: {e}")

def normalize_domain(domain):
    """Возвращает домен портала для REST. oauth.bitrix.info — это OAuth-сервер,
    а не портал: REST-методы там не живут (404 ERROR_METHOD_NOT_FOUND).
    В таком случае подставляем реальный портал."""
    d = (domain or "").strip().strip("/")
    if not d or "oauth.bitrix" in d:
        return BITRIX_PORTAL_DOMAIN
    return d

def _domain_from_endpoint(url):
    """Достаёт хост портала из client_endpoint вида https://joto.bitrix24.ru/rest/."""
    m = re.match(r"https?://([^/]+)", url or "")
    return m.group(1) if m else ""

def get_access_token():
    st = load_oauth()
    if not st:
        return None
    access = st.get("access_token")
    domain = st.get("domain")
    refresh = st.get("refresh_token")
    exp = st.get("expires_at") or 0
    if not access or not domain:
        return None
    if time.time() < exp:
        return access, normalize_domain(domain)
    if not refresh or not BITRIX_CLIENT_ID or not BITRIX_CLIENT_SECRET:
        print("[OAUTH] access_token истёк, refresh невозможен (нет refresh/CLIENT_ID/SECRET)")
        return None
    try:
        r = httpx.get(
            "https://oauth.bitrix.info/oauth/token/",
            params={
                "grant_type": "refresh_token",
                "client_id": BITRIX_CLIENT_ID,
                "client_secret": BITRIX_CLIENT_SECRET,
                "refresh_token": refresh,
            },
            timeout=20,
        )
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        print(f"[OAUTH] refresh failed: {e}")
        return None
    na = data.get("access_token")
    nr = data.get("refresh_token") or refresh
    ein = int(data.get("expires_in") or 3600)
    # Реальный портал берём из client_endpoint, а не из data["domain"] (=oauth.bitrix.info)
    nd = normalize_domain(_domain_from_endpoint(data.get("client_endpoint")) or domain)
    if not na:
        return None
    save_oauth(na, nr, ein, nd, st.get("member_id"))
    print(f"[OAUTH] access_token обновлён через refresh (домен {nd})")
    return na, nd

# ===================== BITRIX REST =====================

def bx_call(method, params=None, auth=None):
    body = dict(params or {})
    if auth and auth.get("access_token") and auth.get("domain"):
        token, domain = auth["access_token"], auth["domain"]
    else:
        tok = get_access_token()
        if not tok:
            raise RuntimeError("Bitrix OAuth не настроен — приложение не установлено?")
        token, domain = tok
    domain = normalize_domain(domain)
    if BITRIX_CLIENT_ID:
        body.setdefault("CLIENT_ID", BITRIX_CLIENT_ID)
    url = f"https://{domain}/rest/{method}?auth={token}"
    r = httpx.post(url, json=body, timeout=30)
    if r.status_code >= 400:
        try:
            b = r.json()
        except Exception:
            b = r.text
        raise RuntimeError(f"Bitrix {method} {r.status_code}: {b}")
    data = r.json()
    if isinstance(data, dict) and data.get("error"):
        raise RuntimeError(f"Bitrix {method} error: {data}")
    return data.get("result") if isinstance(data, dict) else data

def send_b24_message(dialog_id, text, auth=None):
    try:
        st = load_oauth() or {}
        params = {"DIALOG_ID": dialog_id, "MESSAGE": text}
        if st.get("bot_id"):
            params["BOT_ID"] = st["bot_id"]
        bx_call("imbot.message.add", params, auth=auth)
    except Exception as e:
        print(f"Ошибка отправки: {e}")

def register_bot():
    if not EVENT_HANDLER_URL:
        raise RuntimeError("PUBLIC_BASE_URL не задан — некуда слать события")
    result = bx_call("imbot.register", {
        "CODE": BOT_CODE,
        "TYPE": "B",
        "EVENT_MESSAGE_ADD": EVENT_HANDLER_URL,
        "EVENT_WELCOME_MESSAGE": EVENT_HANDLER_URL,
        "EVENT_BOT_DELETE": EVENT_HANDLER_URL,
        "PROPERTIES": {
            "NAME": BOT_NAME,
            "COLOR": "AQUA",
            "EMAIL": "joto-bot@joto.local",
            "WORK_POSITION": "Бот для создания артикулов JOTO",
        },
    })
    bot_id = str(result)
    set_bot_id(bot_id)
    print(f"Бот зарегистрирован: BOT_ID={bot_id}")
    return bot_id

def _bind_left_menu_item(handler_url, title, description):
    """Привязывает один пункт левого меню Bitrix24 (с пере-привязкой без дублей)."""
    if not handler_url:
        raise RuntimeError("PUBLIC_BASE_URL не задан — некуда вести пункт меню")
    try:
        bx_call("placement.unbind", {"PLACEMENT": "LEFT_MENU", "HANDLER": handler_url})
    except Exception:
        pass
    result = bx_call("placement.bind", {
        "PLACEMENT": "LEFT_MENU",
        "HANDLER": handler_url,
        "TITLE": title,
        "DESCRIPTION": description,
    })
    print(f"Пункт левого меню зарегистрирован: {handler_url} ({title})")
    return result

def register_left_menu():
    """Добавляет приложение пунктами в левое меню Bitrix24: карточки WB и распродажа."""
    _bind_left_menu_item(LEFT_MENU_HANDLER_URL, LEFT_MENU_TITLE,
                         "Массовое создание и редактирование карточек Wildberries")
    _bind_left_menu_item(SEASON_MENU_HANDLER_URL, SEASON_MENU_TITLE,
                         "Отчёт по распродаже сезонных товаров (остатки, динамика, скидки)")
    return True

def reset_left_menu():
    """
    Полная пересборка пунктов левого меню.
    Сначала снимает ВСЕ привязки LEFT_MENU (в т.ч. «осиротевшие» после
    переустановки приложения — они вызывают «Приложение не найдено»),
    затем привязывает наши пункты заново.
    """
    removed = 0
    try:
        existing = bx_call("placement.get") or []
        for pl in existing:
            if isinstance(pl, dict) and pl.get("placement") == "LEFT_MENU":
                handler = pl.get("handler")
                if handler:
                    try:
                        bx_call("placement.unbind", {"PLACEMENT": "LEFT_MENU", "HANDLER": handler})
                        removed += 1
                    except Exception as e:
                        print(f"[МЕНЮ] не удалось снять {handler}: {e}")
    except Exception as e:
        print(f"[МЕНЮ] placement.get недоступен: {e}")
    print(f"[МЕНЮ] снято старых привязок LEFT_MENU: {removed}")
    register_left_menu()
    return removed

# ===================== ДИАЛОГ: СОЗДАНИЕ АРТИКУЛА =====================

def send_welcome(dialog_id, auth=None):
    user_states[dialog_id] = {"step": "start"}
    send_b24_message(dialog_id,
        f"👋 Привет! Я бот «{BOT_NAME}» для создания артикулов JOTO.\n\n"
        "Напиши *артикул* чтобы создать новый артикул.\n\n"
        f"Доступные категории:\n{CATS_LIST}",
        auth=auth,
    )

def handle_message(dialog_id, text, auth=None):
    text = text.strip()
    state = user_states.get(dialog_id, {})
    step = state.get("step", "start")

    print(f"handle_message: dialog_id={dialog_id}, step={step}, text={text}")

    if any(word in text.lower() for word in ["помощь", "help", "начать", "старт", "привет", "/start"]):
        send_welcome(dialog_id, auth=auth)
        return

    if text.lower() in ["артикул", "создать", "новый"]:
        user_states[dialog_id] = {"step": "wait_category"}
        send_b24_message(dialog_id,
            f"📦 *Создание артикула*\n\n"
            f"Шаг 1/3: Введите категорию товара:\n{CATS_LIST}",
            auth=auth,
        )
        return

    if step == "wait_category":
        category = text.lower()
        category_code = resolve_category_code(category)
        if not category_code:
            send_b24_message(dialog_id, f"❌ Категория не найдена.\n\nВведите одну из:\n{CATS_LIST}\n\n(новые категории можно добавить в приложении)", auth=auth)
            return
        next_num = peek_next_number(category_code)
        user_states[dialog_id] = {"step": "wait_color", "category": category, "category_code": category_code}
        send_b24_message(dialog_id,
            f"✅ Категория: {category.capitalize()} (J{category_code})\n"
            f"Следующий номер модели: *{next_num}*\n\n"
            f"Шаг 2/3: Введите цвет (например: black, white, grey, navy):",
            auth=auth,
        )
        return

    if step == "wait_color":
        color = text.lower().replace(" ", "")
        user_states[dialog_id]["color"] = color
        user_states[dialog_id]["step"] = "wait_name"
        send_b24_message(dialog_id, f"✅ Цвет: {color}\n\nШаг 3/3: Введите название товара:", auth=auth)
        return

    if step == "wait_name":
        category = state["category"]
        category_code = state["category_code"]
        color = state["color"]
        name = text
        model_number = reserve_next_number(category_code)
        article = f"J{category_code}{model_number}/{color}"
        user_states[dialog_id] = {"step": "start"}
        send_b24_message(dialog_id,
            f"✅ *Артикул создан!*\n\n"
            f"🏷 Артикул: *{article}*\n"
            f"📁 Категория: {category.capitalize()}\n"
            f"🎨 Цвет: {color}\n"
            f"📝 Название: {name}\n"
            f"🔢 Модель №{model_number}\n\n"
            f"Для создания ещё одного напиши *артикул*",
            auth=auth,
        )
        return

    send_b24_message(dialog_id, "Напиши *артикул* чтобы создать новый артикул, или *помощь* для справки.", auth=auth)

# ===================== CTR МОНИТОРИНГ =====================

previous_ctr = {}

def get_wb_ctr():
    try:
        today = datetime.now().date()
        date_from = (today - timedelta(days=2)).strftime("%Y-%m-%d")
        date_to = today.strftime("%Y-%m-%d")

        url = "https://seller-analytics-api.wildberries.ru/api/analytics/v3/sales-funnel/products"
        headers = {"Authorization": WB_API_TOKEN}
        payload = {
            "dateFrom": date_from,
            "dateTo": date_to,
            "limit": 100,
            "offset": 0,
            "orderBy": {"field": "addToCartCount", "mode": "desc"},
            "selectedPeriod": {"begin": date_from, "end": date_to}
        }

        resp = httpx.post(url, headers=headers, json=payload, timeout=30)
        if resp.status_code != 200:
            print(f"WB API ошибка: {resp.text[:300]}")
            return {}

        data = resp.json()
        items = data.get("data", {}).get("products", []) or data.get("products", []) or []

        result = {}
        for item in items:
            nm_id = item.get("nmID") or item.get("nmId")
            name = item.get("vendorCode", str(nm_id))
            views = item.get("openCardCount", 0) or 0
            clicks = item.get("addToCartCount", 0) or 0
            if nm_id and views > 0:
                result[nm_id] = {"ctr": round(clicks / views * 100, 2), "name": name}

        print(f"CTR: получено артикулов {len(result)}")
        return result
    except Exception as e:
        print(f"Ошибка WB API: {e}")
        return {}

def check_ctr():
    global previous_ctr
    print(f"Проверка CTR: {datetime.now()}")
    current = get_wb_ctr()
    if not current:
        return

    alerts = []
    for nm_id, data in current.items():
        ctr = data["ctr"]
        name = data["name"]
        if nm_id in previous_ctr:
            prev_ctr = previous_ctr[nm_id]["ctr"]
            if prev_ctr > 0 and (prev_ctr - ctr) >= 1.0:
                alerts.append(f"⚠️ {name}: CTR снизился с {prev_ctr}% до {ctr}% (−{round(prev_ctr-ctr,2)}%)")

    previous_ctr = current

    if alerts:
        msg = "📉 *Снижение CTR на Wildberries:*\n\n" + "\n".join(alerts)
        send_b24_message(CTR_ALERT_DIALOG, msg)
    else:
        print("Снижений CTR не найдено")

# ===================== ОТЧЁТ ПО ЗАДАЧАМ =====================

def get_users():
    result = bx_call("user.get", {"ACTIVE": True, "USER_TYPE": "employee"})
    users = {}
    if isinstance(result, list):
        for u in result:
            uid = str(u.get("ID", ""))
            name = f"{u.get('NAME', '')} {u.get('LAST_NAME', '')}".strip()
            if uid and name:
                users[uid] = name
    return users

def get_tasks_for_user(user_id):
    today = datetime.now().strftime("%Y-%m-%dT00:00:00+03:00")
    tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%dT00:00:00+03:00")

    done = bx_call("tasks.task.list", {
        "filter": {"RESPONSIBLE_ID": user_id, "STATUS": 5, ">=CLOSED_DATE": today, "<CLOSED_DATE": tomorrow},
        "select": ["ID", "TITLE"]
    })
    overdue = bx_call("tasks.task.list", {
        "filter": {"RESPONSIBLE_ID": user_id, "!STATUS": [4, 5], "<=DEADLINE": today},
        "select": ["ID", "TITLE"]
    })
    in_progress = bx_call("tasks.task.list", {
        "filter": {"RESPONSIBLE_ID": user_id, "STATUS": 3},
        "select": ["ID", "TITLE"]
    })

    def extract_titles(res):
        if isinstance(res, dict):
            return [t.get("title", "") for t in res.get("tasks", [])]
        elif isinstance(res, list):
            return [t.get("title", t.get("TITLE", "")) for t in res]
        return []

    return extract_titles(done), extract_titles(in_progress), extract_titles(overdue)

def generate_report():
    print(f"Генерация отчёта: {datetime.now()}")
    users = get_users()
    if not users:
        send_b24_message(REPORT_USER_ID, "⚠️ Не удалось получить список сотрудников.")
        return

    today_str = datetime.now().strftime("%d.%m.%Y")
    report_lines = [f"📊 *Отчёт по задачам за {today_str}*"]

    for user_id, name in users.items():
        if user_id == REPORT_USER_ID:
            continue

        done, in_progress, overdue = get_tasks_for_user(user_id)
        lines = [f"\n👤 *{name}*"]

        if done:
            lines.append(f"✅ Выполнено ({len(done)}):")
            for t in done[:5]:
                lines.append(f"  • {t}")
            if len(done) > 5:
                lines.append(f"  ...и ещё {len(done)-5}")

        if in_progress:
            lines.append(f"🔄 В работе ({len(in_progress)}):")
            for t in in_progress[:5]:
                lines.append(f"  • {t}")
            if len(in_progress) > 5:
                lines.append(f"  ...и ещё {len(in_progress)-5}")

        if overdue:
            lines.append(f"❌ Просрочено ({len(overdue)}):")
            for t in overdue[:5]:
                lines.append(f"  • {t}")
            if len(overdue) > 5:
                lines.append(f"  ...и ещё {len(overdue)-5}")

        if not done and not in_progress and not overdue:
            lines.append("  — нет активных задач")

        report_lines.extend(lines)

    send_b24_message(REPORT_USER_ID, "\n".join(report_lines))
    print("Отчёт отправлен")

# ===================== FLASK: УСТАНОВКА И СОБЫТИЯ =====================

INSTALL_FINISH_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Установка</title>
<script src="//api.bitrix24.com/api/v1/"></script></head>
<body>
<p>Устанавливаю бота «Article Generator»...</p>
<script>
  BX24.init(function(){ BX24.installFinish(); });
</script>
</body></html>"""

@app.route("/install", methods=["GET", "POST"])
def install():
    vals = request.values
    auth_id = vals.get("AUTH_ID", "")
    refresh_id = vals.get("REFRESH_ID", "")
    expires = vals.get("AUTH_EXPIRES", "3600")
    member_id = vals.get("member_id", "")
    domain = vals.get("DOMAIN", "") or request.args.get("DOMAIN", "")
    app_token = vals.get("application_token", "") or vals.get("APP_SID", "")

    if auth_id and domain:
        init_db()
        save_oauth(auth_id, refresh_id, expires, domain, member_id, app_token=app_token or None)
        try:
            register_bot()
        except Exception as e:
            print(f"Ошибка регистрации бота при установке: {e}")
        try:
            register_left_menu()
        except Exception as e:
            print(f"Ошибка регистрации пункта левого меню при установке: {e}")
    else:
        print("[INSTALL] не пришли AUTH_ID/DOMAIN — проверь настройки приложения")

    return Response(INSTALL_FINISH_HTML, mimetype="text/html")

@app.route("/bitrix/events", methods=["POST"])
def bitrix_events():
    vals = request.values
    event = vals.get("event", "")
    app_token = vals.get("auth[application_token]", "")

    if BITRIX_APP_TOKEN and app_token and app_token != BITRIX_APP_TOKEN:
        return Response("forbidden", status=403)

    auth = {
        "access_token": vals.get("auth[access_token]", ""),
        "domain": vals.get("auth[domain]", ""),
    }
    if auth["access_token"] and auth["domain"]:
        prev = load_oauth() or {}
        save_oauth(
            auth["access_token"],
            vals.get("auth[refresh_token]", "") or prev.get("refresh_token"),
            vals.get("auth[expires_in]", "3600"),
            auth["domain"],
            vals.get("auth[member_id]", "") or prev.get("member_id"),
            app_token=app_token or None,
        )

    if event == "ONIMBOTMESSAGEADD":
        dialog_id = vals.get("data[PARAMS][DIALOG_ID]", "").strip()
        text = vals.get("data[PARAMS][MESSAGE]", "").strip()
        # В «тихих» чатах (отдел продаж, алерты) бот не ведёт диалог по артикулам
        if dialog_id and text and not is_silent_dialog(dialog_id):
            threading.Thread(target=handle_message, args=(dialog_id, text, auth)).start()

    elif event in ("ONIMBOTWELCOMEMESSAGE", "ONIMBOTJOINCHAT"):
        dialog_id = vals.get("data[PARAMS][DIALOG_ID]", "").strip()
        # Не приветствуем артикульным сообщением в чатах для отчётов
        if dialog_id and not is_silent_dialog(dialog_id):
            threading.Thread(target=send_welcome, args=(dialog_id, auth)).start()

    elif event in ("ONIMBOTDELETE", "ONAPPUNINSTALL"):
        print(f"Событие: {event}")

    return jsonify({"ok": True})

# ===================== FLASK: ИНТЕРФЕЙС ПРИЛОЖЕНИЯ =====================

APP_PAGE_HTML = """<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1"><title>Генерация артикулов</title></head>
<body style="font-family:sans-serif;padding:24px"><h1>Генерация артикулов</h1>
<p>Файл интерфейса app_page.html не найден. Проверьте, что он рядом с main.py.</p></body></html>"""

APP_PAGE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_page.html")

def load_app_page():
    try:
        with open(APP_PAGE_PATH, encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        print(f"Не удалось загрузить app_page.html: {e}")
        return APP_PAGE_HTML

def load_named_page(path, fallback="<!doctype html><meta charset='utf-8'><h1>Страница не найдена</h1>"):
    try:
        with open(path, encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        print(f"Не удалось загрузить {path}: {e}")
        return fallback

LOGO_EXTS = ("png", "svg", "jpg", "jpeg", "webp")

def _logo_file():
    """Путь к файлу logo.* рядом с main.py, если загружен."""
    base = os.path.dirname(os.path.abspath(__file__))
    for ext in LOGO_EXTS:
        p = os.path.join(base, "logo." + ext)
        if os.path.exists(p):
            return p
    return None

@app.route("/logo", methods=["GET"])
def logo():
    p = _logo_file()
    if not p:
        return Response("not found", status=404)
    mt = mimetypes.guess_type(p)[0] or "application/octet-stream"
    with open(p, "rb") as f:
        return Response(f.read(), mimetype=mt)

@app.route("/", methods=["GET", "POST"])
def index():
    return Response(load_app_page(), mimetype="text/html")

@app.route("/api/config", methods=["GET"])
def api_config():
    logo_url = LOGO_URL or ("/logo" if _logo_file() else "")
    return jsonify({"ok": True, "logo_url": logo_url})

@app.route("/api/categories", methods=["GET"])
def api_categories():
    return jsonify({"ok": True, "categories": all_categories()})

@app.route("/api/category", methods=["POST"])
def api_add_category():
    data = request.get_json(silent=True) or request.form
    name = (data.get("name", "") or "").strip().lower()
    code = (data.get("code", "") or "").strip()
    ok, result = add_category(name, code or None)
    if not ok:
        return jsonify({"ok": False, "error": result}), 400
    return jsonify({"ok": True, "value": name, "title": name.capitalize(), "code": result})

@app.route("/api/next", methods=["GET"])
def api_next():
    category = (request.args.get("category", "") or "").strip().lower()
    code = resolve_category_code(category)
    if not code:
        return jsonify({"ok": False, "error": "Неизвестная категория"}), 400
    next_num = peek_next_number(code)
    return jsonify({"ok": True, "category_code": code, "next_number": next_num})

@app.route("/api/article", methods=["POST"])
def api_article():
    data = request.get_json(silent=True) or request.form
    category = (data.get("category", "") or "").strip().lower()
    color = (data.get("color", "") or "").strip().lower().replace(" ", "")
    name = (data.get("name", "") or "").strip()
    code = resolve_category_code(category)
    if not code:
        return jsonify({"ok": False, "error": "Неизвестная категория"}), 400
    if not color:
        return jsonify({"ok": False, "error": "Укажите цвет"}), 400
    model_number = reserve_next_number(code)
    article = f"J{code}{model_number}/{color}"
    return jsonify({
        "ok": True,
        "article": article,
        "category": category,
        "category_title": CATEGORY_TITLES.get(category, category.capitalize()),
        "category_code": code,
        "color": color,
        "name": name,
        "model_number": model_number,
    })

# ===================== WB CONTENT API: МАССОВЫЕ КАРТОЧКИ =====================
# Создание и редактирование карточек товаров на Wildberries.
# Требуется WB_API_TOKEN с доступом к категории «Контент».

WB_CONTENT_BASE = "https://content-api.wildberries.ru"
WB_PRICES_BASE = "https://discounts-prices-api.wildberries.ru"  # API цен и скидок (отдельная категория токена)
WB_CARDS_PAGE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cards_page.html")

def wb_content_request(method, path, json_body=None, params=None, timeout=60):
    if not WB_API_TOKEN:
        raise RuntimeError("WB_API_TOKEN не задан — нужен токен с доступом к категории «Контент».")
    url = WB_CONTENT_BASE + path
    r = httpx.request(method, url, headers={"Authorization": WB_API_TOKEN},
                      json=json_body, params=params, timeout=timeout)
    if r.status_code >= 400:
        try:
            detail = r.json()
        except Exception:
            detail = r.text[:500]
        raise RuntimeError(f"WB {path} {r.status_code}: {detail}")
    try:
        return r.json()
    except Exception:
        return {}

def wb_fetch_cards(text_search="", limit=100, max_cards=1000):
    """Загружает карточки продавца (с пагинацией по курсору)."""
    cards = []
    page = max(1, min(int(limit or 100), 100))
    cursor = {"limit": page}
    for _ in range(200):
        flt = {"withPhoto": -1}
        if text_search:
            flt["textSearch"] = text_search
        payload = {"settings": {"cursor": cursor, "filter": flt}}
        data = wb_content_request("POST", "/content/v2/get/cards/list", json_body=payload)
        batch = data.get("cards", []) or []
        cards.extend(batch)
        cur = data.get("cursor", {}) or {}
        total = cur.get("total", 0)
        if total < cursor["limit"] or len(cards) >= max_cards:
            break
        cursor = {"limit": page, "updatedAt": cur.get("updatedAt"), "nmID": cur.get("nmID")}
    return cards[:max_cards]

def simplify_card(c):
    """Облегчённое представление карточки для интерфейса + сырой объект для обновления."""
    barcodes = []
    for s in (c.get("sizes") or []):
        for sku in (s.get("skus") or []):
            barcodes.append(sku)
    return {
        "nmID": c.get("nmID"),
        "imtID": c.get("imtID"),
        "vendorCode": c.get("vendorCode"),
        "brand": c.get("brand"),
        "title": c.get("title"),
        "description": c.get("description"),
        "subjectID": c.get("subjectID"),
        "subjectName": c.get("subjectName"),
        "barcodes": barcodes,
        "photos": len(c.get("photos") or []),
        "characteristics": [
            {"id": ch.get("id"), "name": ch.get("name"), "value": ch.get("value")}
            for ch in (c.get("characteristics") or [])
        ],
        "raw": c,
    }

def build_update_object(c):
    """Полный объект для /content/v2/cards/update — иначе WB сотрёт незаполненные поля."""
    return {
        "nmID": c.get("nmID"),
        "vendorCode": c.get("vendorCode"),
        "brand": c.get("brand"),
        "title": c.get("title"),
        "description": c.get("description"),
        "dimensions": c.get("dimensions") or {},
        "characteristics": c.get("characteristics") or [],
        "sizes": c.get("sizes") or [],
    }

def wb_update_cards(raw_cards):
    objs = [build_update_object(c) for c in raw_cards]
    results = []
    for i in range(0, len(objs), 1000):  # WB принимает до 3000 за запрос, шлём по 1000
        chunk = objs[i:i + 1000]
        results.append(wb_content_request("POST", "/content/v2/cards/update", json_body=chunk))
    return results

def wb_create_cards(items):
    return wb_content_request("POST", "/content/v2/cards/upload", json_body=items)

def wb_generate_barcodes(count):
    data = wb_content_request("POST", "/content/v2/barcodes", json_body={"count": int(count)})
    if isinstance(data, dict):
        d = data.get("data") or {}
        if isinstance(d, dict):
            return d.get("barcodes") or []
        if isinstance(d, list):
            return d
    return []

def wb_search_subjects(name="", limit=200):
    params = {"locale": "ru", "limit": limit}
    if name:
        params["name"] = name
    data = wb_content_request("GET", "/content/v2/object/all", params=params)
    return data.get("data", []) or []

def wb_subject_charcs(subject_id):
    data = wb_content_request("GET", f"/content/v2/object/charcs/{subject_id}", params={"locale": "ru"})
    return data.get("data", []) or []

def apply_bulk_field(raw, field, value, charc_id=None, charc_name=None):
    """Применяет одно изменение к сырой карточке (in-place)."""
    if field == "description":
        raw["description"] = value
    elif field == "title":
        raw["title"] = value
    elif field == "brand":
        raw["brand"] = value
    elif field == "vendorCode":
        raw["vendorCode"] = value
    elif field == "subjectID":
        try:
            raw["subjectID"] = int(value)
        except Exception:
            raw["subjectID"] = value
    elif field == "characteristic":
        chars = raw.get("characteristics") or []
        new_value = value if isinstance(value, list) else [value]
        found = False
        for ch in chars:
            if charc_id is not None and ch.get("id") == charc_id:
                ch["value"] = new_value
                found = True
                break
            if charc_name and ch.get("name") == charc_name:
                ch["value"] = new_value
                found = True
                break
        if not found and (charc_id is not None or charc_name):
            entry = {"value": new_value}
            if charc_id is not None:
                entry["id"] = charc_id
            if charc_name:
                entry["name"] = charc_name
            chars.append(entry)
        raw["characteristics"] = chars
    elif field == "barcode":
        sizes = raw.get("sizes") or []
        for s in sizes:
            skus = s.get("skus") or []
            if value and value not in skus:
                skus.append(value)
            s["skus"] = skus
        raw["sizes"] = sizes
    return raw

@app.route("/cards", methods=["GET", "POST"])
def cards_page():
    return Response(load_named_page(WB_CARDS_PAGE_PATH), mimetype="text/html")

@app.route("/api/wb/cards", methods=["GET"])
def api_wb_cards():
    search = (request.args.get("search", "") or "").strip()
    try:
        max_cards = int(request.args.get("limit", "1000"))
    except Exception:
        max_cards = 1000
    try:
        cards = wb_fetch_cards(text_search=search, limit=100, max_cards=max_cards)
        return jsonify({"ok": True, "count": len(cards),
                        "cards": [simplify_card(c) for c in cards]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/api/wb/bulk-edit", methods=["POST"])
def api_wb_bulk_edit():
    data = request.get_json(silent=True) or {}
    cards = data.get("cards") or []
    field = (data.get("field") or "").strip()
    value = data.get("value")
    charc_id = data.get("charcId")
    charc_name = data.get("charcName")
    if not cards:
        return jsonify({"ok": False, "error": "Не выбрано ни одной карточки"}), 400
    if not field:
        return jsonify({"ok": False, "error": "Не указано, что менять"}), 400
    try:
        for c in cards:
            apply_bulk_field(c, field, value, charc_id, charc_name)
        results = wb_update_cards(cards)
        errors = [r for r in results if isinstance(r, dict) and r.get("error")]
        if errors:
            msg = "; ".join(str(e.get("errorText") or e.get("additionalErrors") or "ошибка WB") for e in errors)
            return jsonify({"ok": False, "error": msg, "details": errors}), 502
        return jsonify({"ok": True, "updated": len(cards)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/api/wb/set-prices", methods=["POST"])
def api_wb_set_prices():
    """Массовое изменение цены через API цен WB (нужна категория токена «Цены и скидки»).
    items: [{nmID, price}] — price в рублях (целое)."""
    if not WB_API_TOKEN:
        return jsonify({"ok": False, "error": "WB_API_TOKEN не задан"}), 400
    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    payload = []
    for it in items:
        try:
            nm = int(it.get("nmID"))
            price = int(round(float(it.get("price"))))
        except (TypeError, ValueError):
            continue
        if nm and price > 0:
            payload.append({"nmID": nm, "price": price})
    if not payload:
        return jsonify({"ok": False, "error": "Нет корректных пар nmID/цена"}), 400
    try:
        r = httpx.post(WB_PRICES_BASE + "/api/v2/upload/task",
                       headers={"Authorization": WB_API_TOKEN},
                       json={"data": payload}, timeout=60)
        if r.status_code >= 400:
            try:
                detail = r.json()
            except Exception:
                detail = r.text[:400]
            err = detail.get("errorText") if isinstance(detail, dict) else detail
            if r.status_code in (401, 403):
                err = "Нет доступа к API цен — у токена нужна категория «Цены и скидки». " + str(err or "")
            return jsonify({"ok": False, "error": f"WB цены {r.status_code}: {err}"}), 502
        body = r.json() if r.headers.get("content-type", "").startswith("application/json") else {}
        if isinstance(body, dict) and body.get("error"):
            return jsonify({"ok": False, "error": body.get("errorText") or "Ошибка WB цен"}), 502
        return jsonify({"ok": True, "updated": len(payload)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/api/wb/create", methods=["POST"])
def api_wb_create():
    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    if not items:
        return jsonify({"ok": False, "error": "Нет карточек для создания"}), 400
    try:
        result = wb_create_cards(items)
        if isinstance(result, dict) and result.get("error"):
            msg = result.get("errorText") or result.get("additionalErrors") or "Ошибка WB"
            return jsonify({"ok": False, "error": str(msg), "details": result}), 502
        return jsonify({"ok": True, "created": len(items), "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/api/wb/barcodes", methods=["POST"])
def api_wb_barcodes():
    data = request.get_json(silent=True) or {}
    try:
        count = int(data.get("count", 1))
    except Exception:
        count = 1
    count = max(1, min(count, 5000))
    try:
        return jsonify({"ok": True, "barcodes": wb_generate_barcodes(count)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/api/wb/subjects", methods=["GET"])
def api_wb_subjects():
    name = (request.args.get("name", "") or "").strip()
    try:
        subjects = wb_search_subjects(name=name, limit=200)
        items = [{"subjectID": s.get("subjectID"), "subjectName": s.get("subjectName"),
                  "parentName": s.get("parentName")} for s in subjects]
        return jsonify({"ok": True, "subjects": items})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/api/wb/charcs", methods=["GET"])
def api_wb_charcs():
    try:
        subject_id = int(request.args.get("subjectId", "0"))
    except Exception:
        subject_id = 0
    if not subject_id:
        return jsonify({"ok": False, "error": "Не указан subjectId"}), 400
    try:
        charcs = wb_subject_charcs(subject_id)
        items = [{"id": c.get("charcID"), "name": c.get("name"), "required": c.get("required"),
                  "unitName": c.get("unitName"), "maxCount": c.get("maxCount"),
                  "charcType": c.get("charcType")} for c in charcs]
        return jsonify({"ok": True, "charcs": items})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

# ===================== WB STATISTICS API: ОТЧЁТ ПО СЕЗОННОЙ РАСПРОДАЖЕ =====================
# Отчёт по распродаже сезонной категории (по умолчанию «Шорты», код 10):
# остаток → динамика продаж → на сколько хватит → какую скидку/темп «вбить»,
# чтобы к концу сезона осталось не более целевого % (по умолчанию 10 %),
# а остальное не легло в неликвид. Нужен WB_API_TOKEN с доступом к «Статистике».

WB_STATS_BASE = "https://statistics-api.wildberries.ru"

# Сезонные категории JOTO: код артикула (J<код>...) + ключевые слова предмета на WB.
SEASONAL_PRESETS = {
    "10": {"title": "Шорты",   "kw": ["шорт"]},
    "02": {"title": "Куртки",  "kw": ["куртк", "пуховик", "парк"]},
    "11": {"title": "Футболки","kw": ["футболк"]},
    "05": {"title": "Худи",    "kw": ["худи", "толстовк"]},
}

# Справочник фактических начальных остатков (приходы с производств), data/initial_stock.json.
INITIAL_STOCK_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "initial_stock.json")

# Перевод цветов RU→EN: накладные на русском, артикулы WB могут быть на латинице.
COLOR_RU_EN = {
    "черный": "black", "чёрный": "black", "белый": "white", "серый": "grey",
    "серый меланж": "grey melange", "меланж": "melange", "темно-серый": "dark grey",
    "тёмно-серый": "dark grey", "светло-серый": "light grey", "голубой": "blue",
    "светло-голубой": "light blue", "синий": "blue", "темно-синий": "navy",
    "тёмно-синий": "navy", "зеленый": "green", "зелёный": "green", "хаки": "khaki",
    "бежевый": "beige", "коричневый": "brown", "красный": "red", "бордовый": "burgundy",
    "розовый": "pink", "желтый": "yellow", "жёлтый": "yellow", "оранжевый": "orange",
    "фиолетовый": "purple", "оливковый": "olive", "молочный": "milk", "кремовый": "cream",
    "песочный": "sand", "графит": "graphite", "мятный": "mint", "бирюзовый": "turquoise",
}
COLOR_EN_RU = {en: ru for ru, en in COLOR_RU_EN.items()}

def _norm_vendor(v):
    """Нормализуем артикул/цвет для сопоставления (регистр, пробелы)."""
    return re.sub(r"\s+", " ", str(v or "").strip()).lower()

def _vendor_variants(vendor):
    """Все варианты ключа артикула для сопоставления: как есть + перевод цвета RU↔EN."""
    v = _norm_vendor(vendor)
    variants = {v}
    if "/" in v:
        base, color = v.split("/", 1)
        base, color = base.strip(), color.strip()
        if color in COLOR_RU_EN:
            variants.add(f"{base}/{COLOR_RU_EN[color]}")
        if color in COLOR_EN_RU:
            variants.add(f"{base}/{COLOR_EN_RU[color]}")
    return variants

def load_initial_stock():
    """Возвращает lookup {ключ -> item}, где ключи включают перевод цвета RU↔EN.
    item = {'vendorCode', 'total', 'sizes', ...}. {} если файла нет."""
    try:
        with open(INITIAL_STOCK_PATH, encoding="utf-8") as fh:
            items = (json.load(fh) or {}).get("items", {})
    except Exception:
        return {}
    lookup = {}
    for key, item in items.items():
        for k in _vendor_variants(item.get("vendorCode") or key):
            lookup.setdefault(k, item)
    return lookup

def lookup_initial(initial_stock, vendor):
    """Ищет начальный остаток по артикулу, перебирая варианты перевода цвета."""
    for k in _vendor_variants(vendor):
        if k in initial_stock:
            return initial_stock[k]
    return {}

# Кэш ответов WB, чтобы не упираться в лимит (429) при повторных «Сформировать».
_WB_CACHE = {}
_WB_CACHE_TTL = 180  # сек

def _wb_cache_get(key):
    v = _WB_CACHE.get(key)
    if v and (time.time() - v[0]) < _WB_CACHE_TTL:
        return v[1]
    return None

def _wb_cache_set(key, data):
    _WB_CACHE[key] = (time.time(), data)

def wb_stats_request(path, params=None, timeout=120, retries=3):
    if not WB_API_TOKEN:
        raise RuntimeError("WB_API_TOKEN не задан — нужен токен с доступом к категории «Статистика».")
    url = WB_STATS_BASE + path
    detail = None
    for attempt in range(retries + 1):
        r = httpx.get(url, headers={"Authorization": WB_API_TOKEN}, params=params, timeout=timeout)
        if r.status_code == 429:  # лимит WB — подождём и повторим
            try:
                detail = r.json()
            except Exception:
                detail = r.text[:500]
            if attempt < retries:
                time.sleep(min(20, 3 * (2 ** attempt)))  # 3, 6, 12 c
                continue
            raise RuntimeError(f"WB stats {path} 429: {detail}")
        if r.status_code >= 400:
            try:
                detail = r.json()
            except Exception:
                detail = r.text[:500]
            raise RuntimeError(f"WB stats {path} {r.status_code}: {detail}")
        try:
            return r.json() or []
        except Exception:
            return []

def fetch_wb_stocks():
    """Актуальный срез остатков по всем складам (с кэшем)."""
    cached = _wb_cache_get("stocks")
    if cached is not None:
        return cached
    data = wb_stats_request("/api/v1/supplier/stocks", params={"dateFrom": "2020-01-01"})
    _wb_cache_set("stocks", data)
    return data

WB_PRICES_BASE = "https://discounts-prices-api.wildberries.ru"

def fetch_wb_prices():
    """Цены и скидки по номенклатурам из WB Prices API.
    Возвращает {nmId: {'price': базовая_цена, 'discount': %, 'discounted': цена_со_скидкой}}.
    Источник надёжнее, чем поле Price в остатках (его WB отдаёт не всегда).
    Тихо возвращает {} если у токена нет доступа к категории «Цены и скидки»."""
    out = {}
    for g in fetch_wb_prices_goods():
        nm = g.get("nmID")
        disc = g.get("discount") or 0
        sizes = g.get("sizes") or []
        price = next((s.get("price") for s in sizes if s.get("price")), None)
        discounted = next((s.get("discountedPrice") for s in sizes if s.get("discountedPrice")), None)
        if nm is not None and price:
            out[nm] = {"price": price, "discount": disc, "discounted": discounted}
    return out

def fetch_wb_prices_goods():
    """Сырой список номенклатур из WB Prices API (с пагинацией, с кэшем). [] если нет доступа."""
    if not WB_API_TOKEN:
        return []
    cached = _wb_cache_get("prices_goods")
    if cached is not None:
        return cached
    goods_all = []
    offset, limit = 0, 1000
    try:
        for _ in range(50):  # до 50 000 номенклатур
            r = httpx.get(WB_PRICES_BASE + "/api/v2/list/goods/filter",
                          headers={"Authorization": WB_API_TOKEN},
                          params={"limit": limit, "offset": offset}, timeout=60)
            if r.status_code >= 400:
                break
            goods = (((r.json() or {}).get("data") or {}).get("listGoods")) or []
            if not goods:
                break
            goods_all.extend(goods)
            if len(goods) < limit:
                break
            offset += limit
    except Exception:
        return goods_all
    _wb_cache_set("prices_goods", goods_all)
    return goods_all

def fetch_wb_prices_by_vendor():
    """Цены/скидки по артикулу (vendorCode): {норм.артикул: {'price','discount','discounted'}}."""
    out = {}
    for g in fetch_wb_prices_goods():
        vc = g.get("vendorCode")
        if not vc:
            continue
        disc = g.get("discount") or 0
        sizes = g.get("sizes") or []
        price = next((s.get("price") for s in sizes if s.get("price")), None)
        discounted = next((s.get("discountedPrice") for s in sizes if s.get("discountedPrice")), None)
        if price:
            out[_norm_vendor(vc)] = {"price": price, "discount": disc, "discounted": discounted}
    return out

def fetch_wb_orders(date_from):
    """Все заказы с указанной даты, с пагинацией по lastChangeDate (flag=0).

    WB отдаёт заказы пачками (до ~80 000 за ответ). Чтобы охватить весь объём
    (все шорты без потерь), идём по курсору lastChangeDate и дедупим по srid.
    """
    cached = _wb_cache_get(f"orders:{date_from}")
    if cached is not None:
        return cached
    collected = {}
    cursor_from = date_from
    for _ in range(60):  # защита от зацикливания
        batch = wb_stats_request("/api/v1/supplier/orders",
                                 params={"dateFrom": cursor_from, "flag": 0})
        if not batch:
            break
        new_count = 0
        max_lc = cursor_from
        for o in batch:
            key = o.get("srid") or f"{o.get('gNumber')}_{o.get('nmId')}_{o.get('barcode')}_{o.get('date')}"
            if key not in collected:
                collected[key] = o
                new_count += 1
            lc = o.get("lastChangeDate") or ""
            if lc > max_lc:
                max_lc = lc
        # больше нет сдвига по времени или новых записей — конец выгрузки
        if new_count == 0 or max_lc == cursor_from:
            break
        cursor_from = max_lc
    result = list(collected.values())
    _wb_cache_set(f"orders:{date_from}", result)
    return result

def fetch_wb_sales(date_from):
    """Продажи (выкупы) с указанной даты, пагинация по lastChangeDate, дедуп по saleID.
    Записи с saleID 'S...' — выкуп, 'R...' — возврат."""
    cached = _wb_cache_get(f"sales:{date_from}")
    if cached is not None:
        return cached
    collected = {}
    cursor_from = date_from
    for _ in range(60):
        batch = wb_stats_request("/api/v1/supplier/sales",
                                 params={"dateFrom": cursor_from, "flag": 0})
        if not batch:
            break
        new_count = 0
        max_lc = cursor_from
        for s in batch:
            key = s.get("saleID") or f"{s.get('srid')}_{s.get('nmId')}_{s.get('date')}"
            if key not in collected:
                collected[key] = s
                new_count += 1
            lc = s.get("lastChangeDate") or ""
            if lc > max_lc:
                max_lc = lc
        if new_count == 0 or max_lc == cursor_from:
            break
        cursor_from = max_lc
    result = list(collected.values())
    _wb_cache_set(f"sales:{date_from}", result)
    return result

def _match_seasonal(rec, category_code, keywords):
    """Запись относится к нужной категории по артикулу J<код>… или по названию предмета."""
    art = (rec.get("supplierArticle") or "").upper().strip()
    subj = (rec.get("subject") or "").lower()
    if category_code and art.startswith(f"J{category_code}"):
        return True
    for kw in (keywords or []):
        if kw and kw in subj:
            return True
    return False

def _parse_wb_date(s):
    try:
        return datetime.fromisoformat((s or "")[:19]).date()
    except Exception:
        return None

def build_seasonal_report(category_code="10", keywords=None, season_end="2026-08-31",
                          target_remain_pct=10.0, lookback_days=28, elasticity=2.0,
                          cost_share=None, min_margin=None,
                          period_start=None, period_end=None):
    """
    Считает по сезонной категории:
      • остаток (quantityFull по всем складам),
      • темп продаж (заказы/день) за последнее окно и динамику к предыдущему окну,
      • на сколько хватит остатка (days of supply) и дату обнуления,
      • прогноз остатка к концу сезона при текущем темпе,
      • требуемый темп и рекомендуемую скидку, чтобы осталось ≤ target_remain_pct,
      • объём, который иначе ляжет в неликвид.
    """
    preset = SEASONAL_PRESETS.get(category_code, {})
    if keywords is None:
        keywords = preset.get("kw", [])
    title = preset.get("title") or f"Категория J{category_code}"

    initial_stock = load_initial_stock()      # фактический начальный остаток (приходы)
    cat_initial_size = {}                      # size -> начальный остаток по категории
    has_initial = False

    today = datetime.now().date()

    def _as_date(v):
        if not v:
            return None
        try:
            return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
        except Exception:
            return None

    # Окно анализа продаж: либо выбранный период (календарь), либо «последние N дней».
    # Окно считаем включительно по обе границы, поэтому число дней = (end - start) + 1.
    ps, pe = _as_date(period_start), _as_date(period_end)
    if ps and pe and pe >= ps:
        rec_start, rec_end = ps, pe
        win = max(1, (rec_end - rec_start).days + 1)
    else:
        rec_end = today
        win = max(1, int(lookback_days))
        rec_start = rec_end - timedelta(days=win - 1)
    start_recent = rec_start
    # предыдущее окно той же длины (для динамики): [start_prev, rec_start - 1 день]
    start_prev = rec_start - timedelta(days=win)

    try:
        season_end_d = datetime.strptime(season_end, "%Y-%m-%d").date()
    except Exception:
        season_end_d = today + timedelta(days=90)
    days_left = max(1, (season_end_d - today).days)
    target_frac = max(0.0, min(float(target_remain_pct) / 100.0, 1.0))
    elasticity = max(0.2, float(elasticity))

    # Маржа → максимально допустимая скидка: цена со скидкой ≥ себестоимость·(1+мин.маржа).
    cs = SEASON_COST_SHARE if cost_share is None else float(cost_share)
    mm = SEASON_MIN_MARGIN if min_margin is None else float(min_margin)
    cs = min(max(cs, 0.0), 0.95)
    mm = max(mm, 0.0)
    max_disc = min(85, max(0, int(round((1 - cs * (1 + mm)) * 100))))

    def _rec_disc_for(s_stock, s_daily, base_disc, target_units):
        """Рекомендуемая скидка под остаток/темп, чтобы к концу сезона осталось
        не больше target_units штук, но не глубже максимума по марже (max_disc).
        target_units — целевой остаток в штуках (10% от начального остатка)."""
        base = int(round(base_disc or 0))
        if s_stock <= 0:
            return min(max_disc, base) if base > max_disc else base
        if s_daily <= 0:                                  # стоит без продаж — агрессивно
            return min(max_disc, max(base + 30, 40))
        s_req = max(0.0, s_stock - target_units) / days_left
        if s_req > s_daily:                               # не успеваем — поднять скидку
            up = (s_req / s_daily - 1) * 100.0
            return min(max_disc, int(round(base + up / elasticity)))
        return min(base, max_disc)                        # темпа хватает

    # % выкупа берём за длинный «отстоявшийся» период (выкупы отстают от заказов),
    # а темп заказов — за выбранное окно. Так коэффициент выкупа стабильный.
    buyout_days = max(win, SEASON_BUYOUT_DAYS)
    buyout_start = today - timedelta(days=buyout_days)
    orders_from = min(start_prev, buyout_start)

    orders = fetch_wb_orders(orders_from.strftime("%Y-%m-%d"))
    stocks = fetch_wb_stocks()
    prices = fetch_wb_prices()   # цены/скидки по nmId (надёжнее поля Price в остатках)
    try:
        sales = fetch_wb_sales(buyout_start.strftime("%Y-%m-%d"))  # выкупы за 60 дн (для % выкупа)
        sales_available = True
    except Exception:
        sales = []            # лимит/нет доступа — отчёт всё равно строим
        sales_available = False  # % выкупа просто не покажем (вместо ложного 0%)

    def _norm_size(v):
        sz = str(v or "").strip()
        return sz if sz and sz != "0" else "б/р"

    # --- заказы по nmId (+ по размерам): окно (темп/динамика) и длинное окно (для % выкупа) ---
    recent, prev = {}, {}
    recent_size, prev_size = {}, {}   # nm -> {size -> кол-во}
    cat_size = {}                     # size -> {"recent":n, "prev":n, "stock":q}
    orders_long = {}                  # nm -> заказов за buyout_days (для % выкупа)
    cat_orders_long = 0
    for o in orders:
        if not _match_seasonal(o, category_code, keywords):
            continue
        if o.get("isCancel"):
            continue
        d = _parse_wb_date(o.get("date"))
        if not d:
            continue
        nm = o.get("nmId")
        # длинное окно для коэффициента выкупа
        if buyout_start <= d <= today:
            orders_long[nm] = orders_long.get(nm, 0) + 1
            cat_orders_long += 1
        # окно анализа (темп и динамика)
        if d > rec_end:
            continue  # за пределами выбранного периода
        sz = _norm_size(o.get("techSize"))
        cat_size.setdefault(sz, {"recent": 0, "prev": 0, "stock": 0})
        if d >= start_recent:
            recent[nm] = recent.get(nm, 0) + 1
            recent_size.setdefault(nm, {})[sz] = recent_size.setdefault(nm, {}).get(sz, 0) + 1
            cat_size[sz]["recent"] += 1
        elif d >= start_prev:
            prev[nm] = prev.get(nm, 0) + 1
            prev_size.setdefault(nm, {})[sz] = prev_size.setdefault(nm, {}).get(sz, 0) + 1
            cat_size[sz]["prev"] += 1

    # --- продажи (выкупы): окно (для показа) и длинное окно (для коэффициента выкупа) ---
    sales_recent = {}   # nm -> выкупов за окно
    sales_long = {}     # nm -> выкупов за buyout_days
    cat_sales_long = 0
    for sl in sales:
        if not _match_seasonal(sl, category_code, keywords):
            continue
        if not str(sl.get("saleID") or "").upper().startswith("S"):
            continue  # только выкупы (не возвраты)
        d = _parse_wb_date(sl.get("date"))
        if not d:
            continue
        nm = sl.get("nmId")
        if buyout_start <= d <= today:
            sales_long[nm] = sales_long.get(nm, 0) + 1
            cat_sales_long += 1
        if start_recent <= d <= rec_end:
            sales_recent[nm] = sales_recent.get(nm, 0) + 1

    # коэффициент выкупа: по категории (стабильный) и по артикулу (если достаточно заказов)
    cat_buyout_frac = (min(1.0, cat_sales_long / cat_orders_long)
                       if (sales_available and cat_orders_long > 0) else 1.0)

    def _buyout_frac(nm):
        of = orders_long.get(nm, 0)
        if of >= 10 and sales_available:
            return min(1.0, sales_long.get(nm, 0) / of)
        return cat_buyout_frac

    # --- остатки по nmId (+ по размерам, сумма по складам) + мета ---
    stock_by_nm, meta_by_nm = {}, {}
    stock_nm_size = {}   # nm -> {size -> остаток}
    for s in stocks:
        if not _match_seasonal(s, category_code, keywords):
            continue
        nm = s.get("nmId")
        qty = s.get("quantityFull")
        if qty is None:
            qty = (s.get("quantity") or 0) + (s.get("inWayToClient") or 0)
        qty = qty or 0
        sz = _norm_size(s.get("techSize"))
        stock_by_nm[nm] = stock_by_nm.get(nm, 0) + qty
        stock_nm_size.setdefault(nm, {})[sz] = stock_nm_size.setdefault(nm, {}).get(sz, 0) + qty
        cat_size.setdefault(sz, {"recent": 0, "prev": 0, "stock": 0})["stock"] += qty
        if nm not in meta_by_nm:
            meta_by_nm[nm] = {
                "vendorCode": s.get("supplierArticle"),
                "subject": s.get("subject"),
                "brand": s.get("brand"),
                "price": s.get("Price"),
                "discount": s.get("Discount") or 0,
                "size": s.get("techSize"),
            }

    rows = []
    nm_ids = set(stock_by_nm) | set(recent) | set(prev) | set(meta_by_nm)
    for nm in nm_ids:
        meta = meta_by_nm.get(nm, {})
        stock = stock_by_nm.get(nm, 0)
        sold_recent = recent.get(nm, 0)           # заказы за окно
        sold_prev = prev.get(nm, 0)
        sales_cnt = sales_recent.get(nm, 0) if sales_available else None  # выкупы за окно (показ)
        buyout_frac_nm = _buyout_frac(nm)                                  # коэффициент за 60 дн
        buyout_pct = round(buyout_frac_nm * 100) if sales_available else None
        daily = sold_recent / win
        daily_prev = sold_prev / win

        if daily_prev > 0:
            trend = round((daily / daily_prev - 1) * 100, 1)
        elif daily > 0:
            trend = 100.0
        else:
            trend = 0.0

        # Чистый темп списания остатка = заказы × коэффициент выкупа (за 60 дн).
        # Если выкуп недоступен — _buyout_frac вернёт 1.0 (берём заказы как есть).
        eff_daily = daily * buyout_frac_nm

        dos = (stock / eff_daily) if eff_daily > 0 else None
        depletion = (today + timedelta(days=int(round(dos)))).isoformat() if dos is not None else None

        proj_sales = eff_daily * days_left
        proj_left = max(0.0, stock - proj_sales)
        proj_left_pct = round(proj_left / stock * 100, 1) if stock > 0 else 0.0

        # фактический начальный остаток по этому артикулу (из приходов), с переводом цвета
        init_item = lookup_initial(initial_stock, meta.get("vendorCode"))
        init_total = init_item.get("total")
        init_sizes = init_item.get("sizes", {}) or {}
        if init_item:
            has_initial = True
        # продано/ушло со склада с момента прихода = начальный − текущий (≥0)
        sold_since = max(0, int(init_total) - int(stock)) if init_total is not None else None

        # Цель «оставить ≤target%» считаем от ФАКТИЧЕСКОГО НАЧАЛЬНОГО остатка (если он есть),
        # иначе — от текущего остатка.
        base_for_target = init_total if init_total is not None else stock
        target_left_units = base_for_target * target_frac
        need_sell = max(0.0, stock - target_left_units)
        required_daily = need_sell / days_left
        deadstock = max(0.0, round(proj_left - target_left_units))  # сверх плана ляжет в неликвид

        # цена/скидка: приоритет — WB Prices API, фолбэк — поля из остатков
        pinfo = prices.get(nm) or {}
        base_price = pinfo.get("price") or meta.get("price") or 0
        cur_disc = pinfo.get("discount") if pinfo.get("discount") is not None else (meta.get("discount") or 0)
        rec_disc = cur_disc
        status = "ok"
        if stock <= 0:
            status = "empty"
        elif eff_daily <= 0:
            status = "stuck"  # есть остаток, но нет чистых продаж
            rec_disc = min(max_disc, max(int(cur_disc) + 30, 40))
        elif required_daily > eff_daily:
            status = "accelerate"
            uplift = (required_daily / eff_daily - 1) * 100.0  # на сколько % поднять темп
            add_pp = uplift / elasticity
            rec_disc = min(max_disc, int(round(cur_disc + add_pp)))
        else:
            status = "ok"  # текущего темпа хватает

        # Цены в рублях: текущая (со скидкой), рекомендованная и минимальная (предел по марже)
        cur_price = pinfo.get("discounted") or (int(round(base_price * (1 - int(cur_disc) / 100.0))) if base_price else None)
        rec_price = int(round(base_price * (1 - int(rec_disc) / 100.0))) if base_price else None
        min_price = int(round(base_price * (1 - max_disc / 100.0))) if base_price else None

        # разбивка по размерам внутри артикула
        size_rows = []
        sizes_seen = set(stock_nm_size.get(nm, {})) | set(recent_size.get(nm, {})) | set(init_sizes)
        for sz in sizes_seen:
            s_stock = stock_nm_size.get(nm, {}).get(sz, 0)
            s_recent = recent_size.get(nm, {}).get(sz, 0)
            s_init = init_sizes.get(sz)
            if s_init is not None:
                cat_initial_size[sz] = cat_initial_size.get(sz, 0) + int(s_init)
            s_daily = s_recent / win
            s_eff = s_daily * buyout_frac_nm       # чистый темп размера с учётом выкупа
            s_dos = int(round(s_stock / s_eff)) if s_eff > 0 else None
            s_proj_left = max(0.0, s_stock - s_eff * days_left)
            s_pct = round(s_proj_left / s_stock * 100, 1) if s_stock > 0 else 0.0
            s_status = "empty" if s_stock <= 0 else ("stuck" if s_eff <= 0 else "ok")
            # цель по размеру — 10% от начального остатка размера (иначе от текущего)
            s_target = (s_init if s_init is not None else s_stock) * target_frac
            s_rec_disc = _rec_disc_for(s_stock, s_eff, cur_disc, s_target)
            s_rec_price = int(round(base_price * (1 - s_rec_disc / 100.0))) if base_price else None
            size_rows.append({
                "size": sz, "stock": int(s_stock), "soldRecent": s_recent,
                "initialStock": int(s_init) if s_init is not None else None,
                "soldSinceStart": (max(0, int(s_init) - int(s_stock)) if s_init is not None else None),
                "dailyRate": round(s_daily, 2), "daysOfSupply": s_dos,
                "projLeftPct": s_pct, "status": s_status,
                "currentDiscount": int(cur_disc),
                "recommendedDiscount": s_rec_disc,
                "currentPrice": cur_price, "recommendedPrice": s_rec_price,
                "minPrice": min_price,
            })
        size_rows.sort(key=lambda x: x["stock"], reverse=True)

        rows.append({
            "nmId": nm,
            "vendorCode": meta.get("vendorCode"),
            "subject": meta.get("subject"),
            "size": meta.get("size"),
            "price": meta.get("price"),
            "initialStock": int(init_total) if init_total is not None else None,
            "soldSinceStart": sold_since,
            "stock": int(stock),
            "soldRecent": sold_recent,
            "salesRecent": sales_cnt,
            "buyoutPct": buyout_pct,
            "dailyRate": round(daily, 2),
            "trendPct": trend,
            "daysOfSupply": int(round(dos)) if dos is not None else None,
            "depletionDate": depletion,
            "projLeft": int(round(proj_left)),
            "projLeftPct": proj_left_pct,
            "requiredDaily": round(required_daily, 2),
            "deadstock": int(deadstock),
            "currentDiscount": int(cur_disc),
            "recommendedDiscount": int(rec_disc),
            "currentPrice": cur_price,
            "recommendedPrice": rec_price,
            "minPrice": min_price,
            "status": status,
            "sizes": size_rows,
        })

    rows.sort(key=lambda r: (r["deadstock"], r["stock"]), reverse=True)

    # --- сводка ---
    total_stock = sum(r["stock"] for r in rows)
    total_initial = sum(r["initialStock"] for r in rows if r.get("initialStock") is not None) if has_initial else None
    total_sold_since = max(0, total_initial - total_stock) if total_initial is not None else None
    total_recent = sum(r["soldRecent"] for r in rows)
    total_sales = sum((r["salesRecent"] or 0) for r in rows) if sales_available else None
    # % выкупа по категории — за длинное окно (стабильный), он же используется в чистом темпе
    total_buyout = round(cat_buyout_frac * 100) if sales_available else None
    total_prev = sum(prev.values())
    cur_daily = total_recent / win
    prev_daily = total_prev / win
    total_trend = round((cur_daily / prev_daily - 1) * 100, 1) if prev_daily > 0 else (100.0 if cur_daily > 0 else 0.0)
    # чистый темп списания по категории = заказы (за окно) × коэффициент выкупа (за 60 дн)
    eff_cur_daily = cur_daily * cat_buyout_frac

    # цель «оставить ≤target%» — от фактического начального остатка (если есть), иначе от текущего
    target_base = total_initial if total_initial is not None else total_stock
    target_left_units = round(target_base * target_frac)
    need_sell = max(0.0, total_stock - target_left_units)
    required_daily = need_sell / days_left
    proj_sales = eff_cur_daily * days_left
    proj_left = max(0.0, round(total_stock - proj_sales))
    proj_left_pct = round(proj_left / total_stock * 100, 1) if total_stock > 0 else 0.0
    total_deadstock = max(0, round(proj_left - target_left_units))
    dos_total = int(round(total_stock / eff_cur_daily)) if eff_cur_daily > 0 else None

    # средневзвешенная текущая скидка и рекомендуемая для всей категории
    if total_stock > 0:
        avg_disc = sum(r["currentDiscount"] * r["stock"] for r in rows) / total_stock
    else:
        avg_disc = 0.0
    rec_disc_total = round(avg_disc)
    rec_disc_raw = round(avg_disc)            # без ограничения по марже — чтобы понять, упёрлись ли
    uplift_total = 0.0
    if eff_cur_daily > 0 and required_daily > eff_cur_daily:
        uplift_total = (required_daily / eff_cur_daily - 1) * 100.0
        rec_disc_raw = int(round(avg_disc + uplift_total / elasticity))
        rec_disc_total = min(max_disc, rec_disc_raw)
    elif eff_cur_daily <= 0 and total_stock > 0:
        rec_disc_raw = max(int(round(avg_disc)) + 30, 40)
        rec_disc_total = min(max_disc, rec_disc_raw)
    # упёрлись в маржу: нужная скидка глубже, чем позволяет минимальная маржа
    margin_limited = rec_disc_raw > max_disc

    # средневзвешенные цены (₽): текущая и рекомендованная под рекомендуемую скидку
    priced = [r for r in rows if r.get("currentPrice")]
    if priced:
        sw = sum(r["stock"] for r in priced) or 1
        avg_cur_price = int(round(sum(r["currentPrice"] * r["stock"] for r in priced) / sw))
        avg_rec_price = int(round(sum((r.get("recommendedPrice") or r["currentPrice"]) * r["stock"] for r in priced) / sw))
        _minp = [r["minPrice"] for r in priced if r.get("minPrice")]
        avg_min_price = int(round(sum(_minp) / len(_minp))) if _minp else None
    else:
        avg_cur_price = avg_rec_price = avg_min_price = None

    target_basis = "начального остатка" if total_initial is not None else "текущего остатка"
    buyout_note = (f" (чистый темп {eff_cur_daily:.1f} с учётом выкупа {total_buyout}%)"
                   if total_buyout is not None else "")
    if total_stock == 0:
        verdict = "Остатков в категории нет — распродавать нечего."
    elif eff_cur_daily <= 0:
        verdict = (f"Продаж за {win} дн. нет, а на складе {total_stock} шт. "
                   f"Без скидки вся партия уйдёт в неликвид. Старт — скидка ~{rec_disc_total} %.")
    elif required_daily <= eff_cur_daily:
        verdict = (f"Идём в графике: при чистом темпе {eff_cur_daily:.1f} шт/день к {season_end} "
                   f"останется ~{proj_left_pct} % — цель ≤ {int(target_remain_pct)} % от {target_basis} "
                   f"({target_left_units} шт) достижима, скидку держим на уровне ~{rec_disc_total} %.")
    else:
        verdict = (f"Не успеваем: чистый темп {eff_cur_daily:.1f} шт/день{(' (заказы '+format(cur_daily, '.1f')+' × выкуп '+str(total_buyout)+'%)') if total_buyout is not None else ''}. "
                   f"Чтобы к {season_end} осталось ≤ {int(target_remain_pct)} % от {target_basis} ({target_left_units} шт), нужно распродать "
                   f"*{int(round(need_sell))} шт* — темп *{required_daily:.1f} шт/день* (+{round(uplift_total)} % к текущему). "
                   f"Иначе в неликвид ляжет ~{total_deadstock} шт. "
                   f"Рекомендуемая средняя скидка ~{rec_disc_total} %.")
    if margin_limited:
        verdict += (f" ⚠️ Скидка упирается в маржу (макс {max_disc} % при себестоимости "
                    f"{int(round(cs*100))} % и мин. марже {int(round(mm*100))} %) — часть остатка "
                    f"в срок без потери маржи не распродать, решение по цене вручную.")

    # --- сценарии «что если поднять скидку» ---
    # Модель: +1 п.п. скидки ≈ +elasticity% к заказам. Считаем, как изменится темп,
    # за сколько распродадим цель (90%) и сколько останется к концу сезона.
    scenarios = []
    if total_stock > 0 and eff_cur_daily > 0:
        base_days_to_target = need_sell / eff_cur_daily if eff_cur_daily > 0 else None
        for delta in [0, 5, 10, 15, 20, 30]:
            total_disc = int(round(avg_disc + delta))
            # не предлагаем скидку глубже, чем позволяет маржа
            if delta > 0 and total_disc > max_disc:
                continue
            # чистый темп при скидке = заказы × эластичность × % выкупа
            new_daily = eff_cur_daily * (1 + elasticity * delta / 100.0)
            if new_daily <= 0:
                continue
            days_to_target = need_sell / new_daily
            s_proj_left = max(0.0, total_stock - new_daily * days_left)
            s_proj_left_pct = round(s_proj_left / total_stock * 100, 1) if total_stock else 0.0
            saved = int(round(base_days_to_target - days_to_target)) if base_days_to_target else 0
            scenarios.append({
                "addDiscount": delta,                      # +п.п. к текущей скидке
                "discount": total_disc,                    # итоговая скидка, %
                "dailyRate": round(new_daily, 2),          # прогноз темпа, шт/день
                "daysToTarget": int(round(days_to_target)),# за сколько распродадим 90%
                "daysSaved": max(0, saved),                # на сколько дней быстрее, чем сейчас
                "selloutDate": (today + timedelta(days=int(round(days_to_target)))).isoformat(),
                "projLeftPct": s_proj_left_pct,            # остаток к концу сезона, %
                "hitsTarget": days_to_target <= days_left, # успеваем ли к концу сезона
            })

    # --- аналитика по размерам (по всей категории) ---
    _SIZE_ORDER = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "L": 4, "XL": 5,
                   "XXL": 6, "2XL": 6, "XXXL": 7, "3XL": 7, "4XL": 8}

    def _size_key(sz):
        u = str(sz).upper()
        if u in _SIZE_ORDER:
            return (0, _SIZE_ORDER[u], 0.0)
        try:
            return (1, 0, float(u.replace(",", ".")))
        except Exception:
            return (2, 0, 0.0)

    size_summary = []
    for sz, d in cat_size.items():
        st = d.get("stock", 0)
        rc = d.get("recent", 0)
        s_daily = rc / win
        s_eff = s_daily * cat_buyout_frac     # чистый темп размера по категории (с учётом выкупа)
        s_dos = int(round(st / s_eff)) if s_eff > 0 else None
        s_proj = max(0.0, st - s_eff * days_left)
        s_pct = round(s_proj / st * 100, 1) if st > 0 else 0.0
        s_status = "empty" if st <= 0 else ("stuck" if s_eff <= 0 else "ok")
        s_init = cat_initial_size.get(sz)
        s_target = (s_init if s_init is not None else st) * target_frac
        s_rec_d = _rec_disc_for(st, s_eff, avg_disc, s_target)
        # средняя базовая цена по категории (из текущей цены и средней скидки)
        avg_base = (avg_cur_price / (1 - avg_disc / 100.0)) if (avg_cur_price and avg_disc < 100) else None
        z_rec_price = int(round(avg_base * (1 - s_rec_d / 100.0))) if avg_base else None
        size_summary.append({
            "size": sz, "stock": int(st), "soldRecent": rc,
            "initialStock": int(s_init) if s_init is not None else None,
            "soldSinceStart": (max(0, int(s_init) - int(st)) if s_init is not None else None),
            "dailyRate": round(s_daily, 2), "daysOfSupply": s_dos,
            "projLeftPct": s_pct, "status": s_status,
            "currentDiscount": int(round(avg_disc)),
            "recommendedDiscount": s_rec_d,
            "currentPrice": avg_cur_price, "recommendedPrice": z_rec_price,
            "minPrice": avg_min_price,
        })
    size_summary.sort(key=lambda x: _size_key(x["size"]))

    return {
        "title": title,
        "categoryCode": category_code,
        "generatedAt": datetime.now().strftime("%d.%m.%Y %H:%M"),
        "seasonEnd": season_end,
        "daysLeft": days_left,
        "lookbackDays": win,
        "periodStart": rec_start.isoformat(),
        "periodEnd": rec_end.isoformat(),
        "targetRemainPct": float(target_remain_pct),
        "elasticity": elasticity,
        "summary": {
            "totalStock": total_stock,
            "initialStock": total_initial,
            "soldSinceStart": total_sold_since,
            "soldRecent": total_recent,
            "salesRecent": total_sales,
            "buyoutPct": total_buyout,
            "currentDaily": round(cur_daily, 2),
            "trendPct": total_trend,
            "daysOfSupply": dos_total,
            "depletionDate": (today + timedelta(days=dos_total)).isoformat() if dos_total else None,
            "requiredDaily": round(required_daily, 2),
            "needSell": int(round(need_sell)),
            "projLeft": proj_left,
            "projLeftPct": proj_left_pct,
            "targetLeftUnits": target_left_units,
            "targetFromInitial": total_initial is not None,
            "deadstock": total_deadstock,
            "currentDiscount": round(avg_disc, 1),
            "recommendedDiscount": rec_disc_total,
            "currentPrice": avg_cur_price,
            "recommendedPrice": avg_rec_price,
            "minPrice": avg_min_price,
            "pricesAvailable": bool(prices) or avg_cur_price is not None,
            "maxDiscountByMargin": max_disc,
            "marginLimited": margin_limited,
            "costSharePct": int(round(cs * 100)),
            "minMarginPct": int(round(mm * 100)),
            "scenarios": scenarios,
            "sizes": size_summary,
            "verdict": verdict,
        },
        "rows": rows,
        "count": len(rows),
    }

@app.route("/season", methods=["GET", "POST"])
def season_page():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "season_page.html")
    return Response(load_named_page(path), mimetype="text/html")

@app.route("/api/wb/season-report", methods=["GET"])
def api_wb_season_report():
    category = (request.args.get("category", "10") or "10").strip()
    season_end = (request.args.get("seasonEnd", "2026-08-31") or "2026-08-31").strip()
    try:
        target_pct = float(request.args.get("targetPct", "10"))
    except Exception:
        target_pct = 10.0
    try:
        lookback = int(request.args.get("lookback", "28"))
    except Exception:
        lookback = 28
    try:
        elasticity = float(request.args.get("elasticity", "2"))
    except Exception:
        elasticity = 2.0
    kw_param = (request.args.get("kw", "") or "").strip()
    keywords = [k.strip().lower() for k in kw_param.split(",") if k.strip()] or None
    cost_share = None
    if request.args.get("costSharePct"):
        try:
            cost_share = float(request.args.get("costSharePct")) / 100.0
        except Exception:
            cost_share = None
    min_margin = None
    if request.args.get("minMarginPct"):
        try:
            min_margin = float(request.args.get("minMarginPct")) / 100.0
        except Exception:
            min_margin = None
    period_start = (request.args.get("periodStart", "") or "").strip() or None
    period_end = (request.args.get("periodEnd", "") or "").strip() or None
    try:
        report = build_seasonal_report(
            category_code=category, keywords=keywords, season_end=season_end,
            target_remain_pct=target_pct, lookback_days=lookback, elasticity=elasticity,
            cost_share=cost_share, min_margin=min_margin,
            period_start=period_start, period_end=period_end,
        )
        return jsonify({"ok": True, "report": report})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

# ----- Отчёт в чат «Отдел продаж» (Битрикс24) -----

def _fmt_date_ru(iso):
    try:
        return datetime.fromisoformat(iso[:19]).strftime("%d.%m")
    except Exception:
        return iso or "—"

def _trend_arrow(p):
    if p is None or p == 0:
        return "→ 0%"
    return ("▲ +" if p > 0 else "▼ ") + f"{abs(p)}%"

def build_seasonal_report_message(rep):
    """Полный отчёт по сезонной распродаже для чата отдела продаж."""
    s = rep["summary"]
    season_url = (PUBLIC_BASE_URL or "https://web-production-d9c0b.up.railway.app") + "/season"
    dos = s.get("daysOfSupply")
    dos_txt = (f"~{dos} дн" + (f" (обнулится {_fmt_date_ru(s['depletionDate'])})" if s.get("depletionDate") else "")) if dos else "∞ (нет продаж)"
    lines = [
        f"📉 *Распродажа сезона — {rep['title']}*",
        f"Отчёт на {rep.get('generatedAt','')} · до конца сезона {rep['daysLeft']} дн (до {rep['seasonEnd']})",
        "",
        f"📦 Остаток: *{s['totalStock']} шт*",
        f"🛒 За {rep['lookbackDays']} дн"
        + (f" ({_fmt_date_ru(rep['periodStart'])}–{_fmt_date_ru(rep['periodEnd'])})" if rep.get('periodStart') else "")
        + f": заказов {s['soldRecent']} шт"
        + (f" · продаж (выкупов) {s['salesRecent']} шт" if s.get('salesRecent') is not None else "")
        + (f" · выкуп {s['buyoutPct']}% (за 60 дн)" if s.get('buyoutPct') is not None else ""),
    ]
    if s.get("initialStock") is not None:
        lines.append(f"🏭 Начальный остаток (факт приход): *{s['initialStock']} шт* · ушло со склада {s['soldSinceStart']} шт")
    lines += [
        f"⚡ Темп: *{s['currentDaily']} шт/день* (динамика {_trend_arrow(s['trendPct'])})",
        f"⏳ Хватит: {dos_txt}",
        f"🎯 Цель ≤{int(rep['targetRemainPct'])}% от {'начального' if s.get('targetFromInitial') else 'текущего'} остатка "
        f"({s['targetLeftUnits']} шт) → распродать *{s.get('needSell','—')} шт*, темп *{s['requiredDaily']} шт/день*",
        f"🧊 В неликвид при текущем темпе: *~{s['deadstock']} шт* (останется {s['projLeftPct']}%)",
        f"🏷 Скидка: сейчас ~{s['currentDiscount']}% → рекомендуем *{s['recommendedDiscount']}%*"
        + (f" · цена ~{s['currentPrice']}₽ → *~{s['recommendedPrice']}₽*"
           + (f" (мин {s['minPrice']}₽)" if s.get("minPrice") else "") if s.get("currentPrice") else ""),
        f"💰 Маржа: макс скидка *{s.get('maxDiscountByMargin','—')}%* "
        f"(себест. {s.get('costSharePct','?')}% · мин. маржа {s.get('minMarginPct','?')}%)"
        + ("  ⚠️ упёрлись в маржу" if s.get("marginLimited") else ""),
        "",
        f"📝 *Вывод:* {s['verdict']}",
    ]
    # Топ позиций, которые сильнее всего рискуют лечь в неликвид
    risky = [r for r in rep.get("rows", []) if r.get("deadstock", 0) > 0][:10]
    if risky:
        lines.append("")
        lines.append("*Что дожимать (топ по неликвиду):*")
        for r in risky:
            disc = f"{r['currentDiscount']}%→{r['recommendedDiscount']}%" if r['recommendedDiscount'] != r['currentDiscount'] else f"{r['currentDiscount']}%"
            price = (f" → цена {r['currentPrice']}₽→{r['recommendedPrice']}₽"
                     if r.get("currentPrice") and r.get("recommendedPrice") and r['recommendedPrice'] != r['currentPrice'] else "")
            lines.append(f"• {r.get('vendorCode') or r.get('nmId')} — остаток {r['stock']}, {r['dailyRate']}/день, в неликвид {r['deadstock']} шт, скидка {disc}{price}")

    # Аналитика по размерам — где залёживается
    sizes = s.get("sizes", [])
    if sizes:
        lines.append("")
        lines.append("📐 *По размерам* (остаток · заказов/день · останется · скидка):")
        for z in sizes:
            flag = " ⚠️" if z["status"] == "stuck" and z["stock"] > 0 else ""
            rec, cur = z.get("recommendedDiscount"), z.get("currentDiscount")
            disc = f"{cur}%→*{rec}%*" if (rec is not None and rec != cur) else (f"{cur}%" if cur is not None else "")
            lines.append(f"• {z['size']}: {z['stock']} шт · {z['dailyRate']}/день · ~{z['projLeftPct']}% · скидка {disc}{flag}")

    # Сценарии «что если поднять скидку» — прогноз темпа и срока распродажи
    scen = [c for c in s.get("scenarios", []) if c.get("addDiscount", 0) > 0]
    if scen:
        lines.append("")
        lines.append("📈 *Если снизить цену (поднять скидку):*")
        for c in scen:
            mark = "✅" if c["hitsTarget"] else "▫️"
            faster = f", это на {c['daysSaved']} дн быстрее" if c.get("daysSaved") else ""
            lines.append(
                f"{mark} скидка {c['discount']}% → ~{c['dailyRate']} шт/день · "
                f"распродадим за ~{c['daysToTarget']} дн (к {_fmt_date_ru(c['selloutDate'])}){faster}, "
                f"остаток к концу сезона ~{c['projLeftPct']}%"
            )
        if any(c.get("hitsTarget") for c in scen):
            lines.append("✅ — успеваем продать цель до конца сезона")
        else:
            lines.append("▫️ — даже при максимальной скидке цель к концу сезона не достигается")

    lines.append("")
    lines.append(f"📊 Полная таблица: {season_url}")
    lines.append(
        "ⓘ Источник: WB Статистика. Темп = заказы после отмён за период "
        f"{_fmt_date_ru(rep.get('periodStart'))}–{_fmt_date_ru(rep.get('periodEnd'))} "
        f"({rep.get('lookbackDays')} дн), они же списывают остаток. "
        "Остаток = quantityFull по всем складам на сейчас."
    )
    return "\n".join(lines)

def find_users_by_name(query):
    """Ищет сотрудников по имени/фамилии через user.get (FIND ищет по ФИО/почте)."""
    q = (query or "").strip()
    if not q:
        return []
    try:
        res = bx_call("user.get", {"FIND": q, "ACTIVE": True})
    except Exception as e:
        print(f"find_users_by_name: {e}")
        return []
    matches = []
    if isinstance(res, list):
        for u in res:
            uid = str(u.get("ID", "")).strip()
            name = f"{u.get('NAME','')} {u.get('LAST_NAME','')}".strip()
            if uid:
                matches.append({"id": uid, "name": name or uid})
    return matches

def resolve_report_recipient(target=None):
    """Превращает получателя в DIALOG_ID. Возвращает (dialog_id, matches).
    Принимает chatXXXX / ID пользователя / имя сотрудника."""
    t = (target if target is not None else SEASON_REPORT_TO or "").strip()
    if not t:
        return SALES_DEPT_DIALOG, None
    if t.lower().startswith("chat") or t.isdigit():
        return t, None
    matches = find_users_by_name(t)
    if matches:
        return matches[0]["id"], matches
    # имя не нашли — не молчим, шлём резервно на REPORT_USER_ID
    return REPORT_USER_ID, []

def send_seasonal_report(category_code=None, season_end=None, dialog_id=None):
    """Строит отчёт по сезонной категории и шлёт его получателю (по умолчанию — Татьяне)."""
    category_code = category_code or SEASON_REPORT_CATEGORY
    season_end = season_end or SEASON_END_DATE
    if dialog_id:
        dialog = dialog_id
    else:
        dialog, _ = resolve_report_recipient()
    rep = build_seasonal_report(category_code=category_code, season_end=season_end)
    msg = build_seasonal_report_message(rep)
    send_b24_message(dialog, msg)
    print(f"Сезонный отчёт отправлен получателю {dialog} (категория {category_code})")
    return rep, dialog

@app.route("/api/wb/season-report/send", methods=["POST"])
def api_wb_season_report_send():
    data = request.get_json(silent=True) or {}
    category = (data.get("category") or SEASON_REPORT_CATEGORY).strip()
    season_end = (data.get("seasonEnd") or SEASON_END_DATE).strip()
    dialog = (data.get("dialog") or "").strip()
    if not dialog:
        dialog, _ = resolve_report_recipient()
    try:
        try:
            target_pct = float(data.get("targetPct", 10))
        except Exception:
            target_pct = 10.0
        try:
            lookback = int(data.get("lookback", 28))
        except Exception:
            lookback = 28
        cost_share = (float(data["costSharePct"]) / 100.0) if data.get("costSharePct") not in (None, "") else None
        min_margin = (float(data["minMarginPct"]) / 100.0) if data.get("minMarginPct") not in (None, "") else None
        period_start = (data.get("periodStart") or "").strip() or None
        period_end = (data.get("periodEnd") or "").strip() or None
        rep = build_seasonal_report(category_code=category, season_end=season_end,
                                    target_remain_pct=target_pct, lookback_days=lookback,
                                    cost_share=cost_share, min_margin=min_margin,
                                    period_start=period_start, period_end=period_end)
        send_b24_message(dialog, build_seasonal_report_message(rep))
        return jsonify({"ok": True, "dialog": dialog, "count": rep.get("count", 0)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/season-report-now", methods=["GET"])
def season_report_now():
    threading.Thread(target=send_seasonal_report).start()
    return jsonify({"ok": True, "message": "Сезонный отчёт отправляется в отдел продаж"})

@app.route("/api/wb/token-check", methods=["GET"])
def api_wb_token_check():
    """Проверка нового токена: какие категории WB доступны (через /ping каждого API).
    Только чтение, ничего не отправляет и не меняет."""
    if not WB_API_TOKEN:
        return jsonify({"ok": False, "error": "WB_API_TOKEN не задан"}), 400

    def _ping(base):
        try:
            r = httpx.get(base + "/ping", headers={"Authorization": WB_API_TOKEN}, timeout=15)
            return {"ok": r.status_code == 200, "httpStatus": r.status_code}
        except Exception as e:
            return {"ok": False, "error": str(e)[:200]}

    cats = {
        "statistics": _ping("https://statistics-api.wildberries.ru"),  # остатки, заказы
        "prices": _ping("https://discounts-prices-api.wildberries.ru"),  # цены и скидки
        "content": _ping("https://content-api.wildberries.ru"),         # карточки
    }
    # Проверка ЗАПИСИ в Контент: безопасный пустой cards/update ([]), ловим read-only
    write = None
    write_reason = ""
    try:
        rw = httpx.post("https://content-api.wildberries.ru/content/v2/cards/update",
                        headers={"Authorization": WB_API_TOKEN}, json=[], timeout=20)
        body = (rw.text or "").lower()
        if rw.status_code == 401 and "read-only" in body:
            write = False
            write_reason = "Токен «только на чтение» — редактирование карточек запрещено"
        elif rw.status_code in (200, 400):
            write = True
        else:
            write = None
            write_reason = f"HTTP {rw.status_code}: {(rw.text or '')[:160]}"
    except Exception as e:
        write_reason = str(e)[:160]
    if write is True:
        verdict = "✅ Токен с записью — редактирование карточек будет работать"
    elif write is False:
        verdict = "❌ Токен только на чтение — создайте токен «Контент» БЕЗ галочки «Только на чтение»"
    else:
        verdict = "⚠ Запись не определена: " + write_reason
    return jsonify({
        "ok": all(v.get("ok") for v in cats.values()),
        "categories": cats,
        "content_write": write,
        "verdict": verdict,
        "hint": "statistics — распродажа (остатки/заказы), prices — цены, content — карточки",
    })

PRICE_LIST_SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "price_list_source.json")

@app.route("/api/wb/price-list.xlsx", methods=["GET"])
def api_wb_price_list():
    """Отдаёт загруженную таблицу (Артикул/Наименование/Штрих-код/Остаток)
    с подтянутыми из WB ценами: Цена, Скидка %, Цена со скидкой. Excel-файл."""
    try:
        from openpyxl import Workbook
        from io import BytesIO
    except Exception:
        return jsonify({"ok": False, "error": "openpyxl не установлен на сервере"}), 500
    try:
        with open(PRICE_LIST_SRC, encoding="utf-8") as fh:
            src = json.load(fh)
    except Exception as e:
        return jsonify({"ok": False, "error": f"нет исходной таблицы: {e}"}), 404

    prices = fetch_wb_prices_by_vendor()  # {норм.артикул: {price, discount, discounted}}

    def _price_for(article):
        for k in _vendor_variants(article):
            if k in prices:
                return prices[k]
        return {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Цены"
    ws.append(["Артикул", "Наименование", "Штрих-код", "Остаток",
               "Цена, ₽", "Скидка, %", "Цена со скидкой, ₽"])
    matched = 0
    for row in src.get("rows", []):
        p = _price_for(row.get("article", ""))
        if p:
            matched += 1
        base = p.get("price")
        disc = p.get("discount")
        discounted = p.get("discounted") or (
            int(round(base * (1 - (disc or 0) / 100.0))) if base else None)
        ws.append([row.get("article"), row.get("name"), row.get("barcode"), row.get("stock"),
                   base, disc, discounted])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"wb_prices_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        bio.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"',
                 "X-Matched": str(matched)},
    )

@app.route("/season-report-debug", methods=["GET"])
def season_report_debug():
    """Синхронная диагностика: показывает, где обрывается путь отчёта в чат."""
    out = {
        "wb_token_set": bool(WB_API_TOKEN),
        "recipient_setting": SEASON_REPORT_TO,
        "category": SEASON_REPORT_CATEGORY,
        "season_end": SEASON_END_DATE,
    }
    # 1) Доступ к статистике WB
    try:
        stocks = fetch_wb_stocks()
        out["wb_stocks_total"] = len(stocks)
    except Exception as e:
        out["step"] = "wb_stocks_failed"
        out["error"] = str(e)[:400]
        return jsonify(out)
    # 2) Построение отчёта по шортам
    try:
        rep = build_seasonal_report(category_code=SEASON_REPORT_CATEGORY, season_end=SEASON_END_DATE)
        out["shorts_positions"] = rep.get("count", 0)
        out["shorts_total_stock"] = rep["summary"]["totalStock"]
        out["shorts_sold_recent"] = rep["summary"]["soldRecent"]
    except Exception as e:
        out["step"] = "build_report_failed"
        out["error"] = str(e)[:400]
        return jsonify(out)
    # 3) Состояние OAuth/бота Битрикса
    st = load_oauth() or {}
    out["oauth_present"] = bool(st.get("access_token"))
    out["bot_id"] = st.get("bot_id")
    out["bitrix_domain"] = st.get("domain")
    if not st.get("access_token"):
        out["step"] = "bitrix_oauth_missing"
        out["error"] = "Приложение Битрикса не установлено или токен не сохранён — некому слать сообщение."
        return jsonify(out)
    # 4) Кому уходит отчёт (резолвим имя получателя в DIALOG_ID)
    recipient, matches = resolve_report_recipient()
    out["recipient_dialog"] = recipient
    if matches is not None:
        out["recipient_matches"] = matches
    # 5) Реальная отправка (ошибку показываем, а не глотаем)
    try:
        params = {"DIALOG_ID": recipient, "MESSAGE": build_seasonal_report_message(rep)}
        if st.get("bot_id"):
            params["BOT_ID"] = st["bot_id"]
        res = bx_call("imbot.message.add", params)
        out["step"] = "sent"
        out["sent"] = True
        out["message_id"] = res
    except Exception as e:
        out["step"] = "send_failed"
        out["sent"] = False
        out["error"] = str(e)[:500]
    return jsonify(out)

# ===================== НАЦИОНАЛЬНЫЙ КАТАЛОГ (ЧЕСТНЫЙ ЗНАК): ГТИНЫ =====================
# Каркас Варианта Б: создать товары в НК → получить ГТИНы → подставить в карточки WB.

@app.route("/api/nk/status", methods=["GET"])
def api_nk_status():
    return jsonify({"ok": True, **nk.nk_status()})

@app.route("/api/nk/create-gtins", methods=["POST"])
def api_nk_create_gtins():
    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    if not items:
        return jsonify({"ok": False, "error": "Нет товаров для создания ГТИНов"}), 400
    try:
        gtin_map = nk.create_gtins_for_items(items)
        return jsonify({"ok": True, "gtins": gtin_map, "count": len(gtin_map)})
    except nk.NKNotConfigured as e:
        return jsonify({"ok": False, "error": str(e), "needs_config": True}), 400
    except nk.NKError as e:
        return jsonify({"ok": False, "error": str(e)}), 502
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.route("/api/wb/parse-nk-file", methods=["POST"])
def api_parse_nk_file():
    """Разбирает Excel-шаблон Национального каталога (Честный ЗНАК):
    достаёт по каждой строке артикул (Модель / артикул производителя) и
    присвоенный ГТИН (Код товара). Возвращает строки для заполнения таблицы."""
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "Файл не загружен"}), 400
    try:
        import openpyxl
    except Exception:
        return jsonify({"ok": False, "error": "На сервере не установлен openpyxl (передеплойте после обновления requirements.txt)"}), 500
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = None
        for s in wb.worksheets:
            if s.title.lower().startswith("import"):
                ws = s
                break
        ws = ws or wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            return jsonify({"ok": False, "error": "В файле нет данных (ожидается шаблон импорта НК)"}), 400

        title_row = rows[0]
        code_row = rows[1] if len(rows) > 1 else ()
        marker_row = rows[2] if len(rows) > 2 else ()
        ncols = max(len(title_row), len(code_row))

        def cell(row, j):
            return (str(row[j]).strip() if row is not None and j < len(row) and row[j] is not None else "")

        def find_col(pred):
            for j in range(ncols):
                if pred(cell(title_row, j), cell(code_row, j), cell(marker_row, j)):
                    return j
            return -1

        import re
        gtin_col = find_col(lambda t, c, m: c.upper() == "GTIN" or t.lower() == "код товара")
        article_col = find_col(lambda t, c, m: "артикул производител" in t.lower() and m.lower() == "value")
        if article_col < 0:
            article_col = find_col(lambda t, c, m: "артикул" in t.lower() and m.lower() == "value")
        name_col = find_col(lambda t, c, m: "полное наименование" in t.lower())
        color_col = find_col(lambda t, c, m: t.lower() == "цвет")
        size_col = find_col(lambda t, c, m: "размер" in t.lower() and m.lower() == "value")
        result_col = find_col(lambda t, c, m: "результат обработки" in t.lower())
        if article_col < 0:
            return jsonify({"ok": False, "error": "Не найден столбец «Модель / артикул производителя» — это шаблон импорта НК?"}), 400

        def extract_gtin(r):
            # 1) прямой столбец «Код товара»/GTIN
            g = cell(r, gtin_col) if gtin_col >= 0 else ""
            g = re.sub(r"\D", "", g)
            if g:
                return g
            # 2) из текста «Результат обработки»: «Создан код товара 4640515801332 ...»
            if result_col >= 0:
                m = re.search(r"код товара\s*(\d{8,14})", cell(r, result_col), re.IGNORECASE)
                if not m:
                    m = re.search(r"(\d{12,14})", cell(r, result_col))
                if m:
                    return m.group(1)
            return ""

        out = []
        for r in rows[4:]:  # данные начинаются с 5-й строки
            art = cell(r, article_col)
            if not art:
                continue
            out.append({
                "article": art,
                "size": cell(r, size_col) if size_col >= 0 else "",
                "gtin": extract_gtin(r),
                "name": cell(r, name_col) if name_col >= 0 else "",
                "color": cell(r, color_col) if color_col >= 0 else "",
            })
        with_gtin = sum(1 for x in out if x["gtin"])
        articles = sorted(set(x["article"] for x in out))
        return jsonify({
            "ok": True, "rows": out, "count": len(out), "with_gtin": with_gtin,
            "articles": len(articles),
            "columns": {"article": article_col + 1,
                        "gtin": (gtin_col + 1) if gtin_col >= 0 else None,
                        "size": (size_col + 1) if size_col >= 0 else None,
                        "result": (result_col + 1) if result_col >= 0 else None,
                        "name": (name_col + 1) if name_col >= 0 else None},
        })
    except Exception as e:
        return jsonify({"ok": False, "error": "Не удалось разобрать файл: " + str(e)}), 500

# ===================== FLASK: СЕРВИСНЫЕ ЭНДПОИНТЫ =====================

@app.route("/admin/bitrix/register", methods=["GET"])
def admin_register():
    if not BITRIX_CLIENT_SECRET or request.args.get("secret", "") != BITRIX_CLIENT_SECRET:
        return Response("forbidden", status=403)
    try:
        bot_id = register_bot()
        return jsonify({"ok": True, "bot_id": bot_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/admin/bitrix/placement", methods=["GET"])
def admin_placement():
    if not BITRIX_CLIENT_SECRET or request.args.get("secret", "") != BITRIX_CLIENT_SECRET:
        return Response("forbidden", status=403)
    try:
        removed = reset_left_menu()
        return jsonify({"ok": True, "removed_old": removed, "items": [
            {"handler": LEFT_MENU_HANDLER_URL, "title": LEFT_MENU_TITLE},
            {"handler": SEASON_MENU_HANDLER_URL, "title": SEASON_MENU_TITLE},
        ]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/wb-refresh", methods=["GET"])
def api_wb_refresh():
    _wb_used_cache["ts"] = 0.0
    try:
        codes = fetch_wb_vendor_codes()
        _wb_used_cache["used"] = _used_numbers_by_code(codes)
        _wb_used_cache["ts"] = time.time()
        summary = {cc: len(s) for cc, s in _wb_used_cache["used"].items()}
        return jsonify({"ok": True, "total_codes": len(codes), "by_category": summary})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/check-now", methods=["GET"])
def check_now():
    threading.Thread(target=check_ctr).start()
    return jsonify({"ok": True})

@app.route("/report-now", methods=["GET"])
def report_now():
    threading.Thread(target=generate_report).start()
    return jsonify({"ok": True, "message": "Отчёт генерируется"})

# ===================== ЗАПУСК =====================

def run_scheduler():
    schedule.every().day.at("06:00").do(check_ctr)        # 09:00 МСК
    schedule.every().day.at("15:00").do(generate_report)  # 18:00 МСК
    schedule.every().day.at("06:00").do(send_seasonal_report)  # 09:00 МСК, ежедневно
    print("Планировщик запущен:")
    print("  - CTR проверка каждый день в 09:00 МСК")
    print("  - Отчёт по задачам каждый день в 18:00 МСК")
    print("  - Отчёт по сезонной распродаже каждый день в 09:00 МСК")
    while True:
        schedule.run_pending()
        time.sleep(60)

def ensure_left_menu_on_start():
    """После деплоя привязывает ТОЛЬКО отсутствующие пункты левого меню,
    не трогая уже привязанные — чтобы их позиция в меню (заданная
    перетаскиванием) не сбрасывалась при каждом перезапуске."""
    if not load_oauth():
        print("[МЕНЮ] OAuth ещё не настроен — пункт левого меню привяжется при установке")
        return
    items = [
        (LEFT_MENU_HANDLER_URL, LEFT_MENU_TITLE,
         "Массовое создание и редактирование карточек Wildberries"),
        (SEASON_MENU_HANDLER_URL, SEASON_MENU_TITLE,
         "Отчёт по распродаже сезонных товаров (остатки, динамика, скидки)"),
    ]
    try:
        existing = set()
        try:
            for pl in (bx_call("placement.get") or []):
                if isinstance(pl, dict) and pl.get("placement") == "LEFT_MENU" and pl.get("handler"):
                    existing.add(pl["handler"].rstrip("/"))
        except Exception as e:
            print(f"[МЕНЮ] placement.get недоступен: {e}")
        for handler, title, desc in items:
            if not handler:
                continue
            if handler.rstrip("/") in existing:
                print(f"[МЕНЮ] уже привязан, позиция сохранена: {handler}")
                continue
            try:
                bx_call("placement.bind", {
                    "PLACEMENT": "LEFT_MENU", "HANDLER": handler,
                    "TITLE": title, "DESCRIPTION": desc,
                })
                print(f"[МЕНЮ] добавлен пункт: {handler} ({title})")
            except Exception as e:
                print(f"[МЕНЮ] не удалось привязать {handler}: {e}")
    except Exception as e:
        print(f"[МЕНЮ] Не удалось проверить пункты левого меню при старте: {e}")

if __name__ == "__main__":
    init_db()
    threading.Thread(target=run_scheduler, daemon=True).start()
    threading.Thread(target=ensure_left_menu_on_start, daemon=True).start()
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
